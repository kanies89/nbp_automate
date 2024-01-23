import shutil
import sys
import re
import os

import datetime
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl.utils as utils
import openpyxl

from tqdm import tqdm
import time
import multiprocessing as mp

from connect import connect, connect_single_query
from variables import EXCEL_READ_AR2, AR2_4_row_1, AR2_4_row_2, EXCEL_READ_AR1, COUNTRY_DICT, countries_to_filter
from f_visa import f_visa_make, check_quarter, read_remote_file
from f_mastercard import f_mastercard_make
from PyQt5.QtCore import pyqtSignal, QObject
from check_rules import to_float

bug_table = []

TEMP = f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/'
PATH = 'Example/'


def reference(index, type_ref):
    """
    Change cell address from DataFrame to Openpyxl. For example df.iat[0,0] == wb["A1"]

    :param index: Value of an index - column or row, depending on type_ref param
    :param type_ref: Choose if you need column or row value translation
    :return: Column or Row value for openpyxl format.
    """

    if type_ref == "col":
        reference_address = utils.get_column_letter(index + 1)
    elif type_ref == "row":
        reference_address = index + 1
    else:
        reference_address = None

    return f'{reference_address}'


class Logger(QObject):
    # Define a custom signal for log updates
    log_updated = pyqtSignal(str)

    def __init__(self, log_file):
        super(Logger, self).__init__()
        self.terminal = sys.stdout
        self.log_file = log_file

    def write(self, message):
        self.terminal.write(message)
        self.log_file.write(message)
        self.log_updated.emit(message)

    def flush(self):
        self.terminal.flush()
        self.log_file.flush()


class TqdmExtraFormat(tqdm):
    """Provides a `minutes per iteration` format parameter"""

    @property
    def format_dict(self):
        d = super(TqdmExtraFormat, self).format_dict
        rate_min = '{:.2f}'.format(1 / d["rate"] / 60) if d["rate"] else '?'
        d.update(rate_min=(rate_min + ' min/' + d['unit']))
        return d


def create_folder_structure(address):
    # Check if the folder structure exists
    if not os.path.exists(address):
        # Create the folder structure
        os.makedirs(address)
        print("Folder structure created successfully.")
    else:
        print("Folder structure already exists.")


def update_bar(queue, total):
    pbar = tqdm(total=round(total, 2))

    while True:
        if pbar.n >= pbar.total:
            break
        time.sleep(0.1)  # Simulate some processing time
        pbar.update(1)
        elapsed_time = (time.time() - pbar.start_t)
        pbar.set_postfix(elapsed=f"{elapsed_time:.2f}s",
                         remaining=f"{(pbar.total - pbar.n) * (elapsed_time / pbar.n):.2f}s / {((pbar.total - pbar.n) * (elapsed_time / pbar.n)) / 60:.2f}min")

    pbar.close()


def aggr_country(c):
    if c == 'PL':
        return 'PL'
    else:
        return 'NPL'


def measure_time_with_progress(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()  # Start time
        result = func(*args, **kwargs)
        elapsed_time = time.time() - start_time  # Elapsed time
        with open("time.txt", "w") as file:
            file.write(str(elapsed_time))  # Save elapsed time to file
        print(f"\n Elapsed time: {elapsed_time:.2f}s")
        return result

    return wrapper


def progress_bar_with_elapsed_time(data):
    with tqdm(total=len(data), ncols=80, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}') as pbar:
        for _ in data:
            time.sleep(0.1)  # Simulate some processing time
            pbar.update(1)


def to_log():
    report_date = datetime.datetime.now().strftime("%Y-%m-%d")
    file_name = f'Log/{report_date}_LOG.txt'
    with open(file_name, 'a') as log:
        log.write(f"\nReport start -AR2- --{report_date}--\n")
    return file_name


def load_df(length, path):
    df = []
    for n in range(length):
        df.append(pd.read_csv(f'{path}{n}.csv', keep_default_na=False))
    return df


def load_or_query(length, name, temp_table, query):
    for i in range(length):
        if os.path.exists(f'{TEMP}{name}{i}.csv'):
            if i == (length - 1) and os.path.exists(f'{TEMP}{name}{i}.csv'):
                dataframe = load_df(length, f'{TEMP}{name}')
                break
            continue
        else:
            dataframe = connect(temp_table, query)
            i = 0
            for df in dataframe:
                df.to_csv(f'{TEMP}{name}{i}.csv')
                i += 1
            break
    return dataframe


def prepare_data_ar2(user, passw, wb_sheets, progress_callback=None, progress_callback_text=None):
    # 4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW
    global visa, mastercard, df_nbp_2

    if progress_callback_text:
        progress_callback_text(f'AR2 - Query: 4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW.')

    print('4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW')

    temp_table = f"Query\\AR2\\NBP_Temp_1.sql"
    query = f"Query\\AR2\\NBP_Query_1.sql"

    dataframe_1 = load_or_query(1, '4.a.R.L_PLiW2_4a.R.W_PLiW2_6.ab.LiW__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(15)

    # Remove spaces from country code
    dataframe_1[0]['country'] = dataframe_1[0]['country'].str.replace(' ', '')

    # Function to replace values based on condition
    def replace_country(value):
        if len(value) > 2:
            value = COUNTRY_DICT.get(value, value)
            if value == "QZ":
                return "D09"
            elif pd.isnull(value):
                return "PL"
            else:
                return value
        else:
            return value

    def replace_card_type(value):
        if pd.isnull(value):
            return "Debit"
        else:
            return value

    def add_g1_countries(value):
        if value in countries_to_filter:
            return value
        else:
            value = "G1"
            return value

    # Apply the function to the first column
    dataframe_1[0]['country'] = dataframe_1[0]['country'].apply(replace_country)

    # Apply the function to the typ_karty column
    dataframe_1[0]['Typ_karty'] = dataframe_1[0]['Typ_karty'].apply(replace_card_type)

    df_for_mcc = dataframe_1[0].copy()

    # Columns:
    # (0) karta / (1) country / (2) mcc / (3) Typ_karty / (4) te_pos_entry_mode / (5) czy_moto /
    # (6) czy_niskokwotowa_zblizeniowa / (7) czy_SCA / (8) ilosc / (9) wartosc_transakcji

    # Filter the DataFrame based on the list of countries
    dataframe_1[0]['country'] = dataframe_1[0]['country'].apply(add_g1_countries)
    filtered_df = dataframe_1[0].copy()

    # List of conditions
    conditions = [
        # 0
        (filtered_df['czy_moto'] == "MOTO"),  # 1
        (filtered_df['czy_moto'] == "MOTO"),  # 2
        (filtered_df['czy_moto'] != "MOTO"),  # 3
        (filtered_df['czy_moto'] != "MOTO"),  # 4
        (filtered_df['czy_moto'] != "MOTO"),  # 5
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "VISA"),  # 6
        (filtered_df['czy_moto'] != "MOTO") & (
                filtered_df['karta'] == "VISA") & (filtered_df['Typ_karty'] == "debit"),  # 7
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "VISA") & (
                (filtered_df['Typ_karty'] != 'debit') & (
                filtered_df['Typ_karty'] != 'credit') | pd.isnull(filtered_df['Typ_karty'])),  # 8
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "VISA") & (
                filtered_df['Typ_karty'] != 'credit'),  # 9
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "VISA") & (filtered_df['czy_SCA'] == 1),  # 10
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "VISA") & (filtered_df['czy_SCA'] == 0),  # 11
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "MC"),  # 12
        (filtered_df['czy_moto'] != "MOTO") & (
                filtered_df['karta'] == "MC") & (filtered_df['Typ_karty'] == "debit"),  # 13
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "MC") & (
                (filtered_df['Typ_karty'] != 'debit') & (
                filtered_df['Typ_karty'] != 'credit') | pd.isnull(filtered_df['Typ_karty'])),  # 14
        (filtered_df['czy_moto'] != "MOTO") & (
                filtered_df['karta'] == "MC") & (filtered_df['Typ_karty'] == "credit"),  # 15
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "MC") & (
                filtered_df['Typ_karty'] == "credit") & (filtered_df['czy_SCA'] == 1),  # 16
        (filtered_df['czy_moto'] != "MOTO") & (filtered_df['karta'] == "MC") & (
                filtered_df['Typ_karty'] == "credit") & (filtered_df['czy_SCA'] == 0),  # 17
        (filtered_df['czy_moto'] != "MOTO") & (
                filtered_df['czy_niskokwotowa_zblizeniowa'] == 1) & (filtered_df['czy_SCA'] == 0),  # 18
        (filtered_df['czy_moto'] != "MOTO") & (
                filtered_df['czy_niskokwotowa_zblizeniowa'] == 0) & (filtered_df['czy_SCA'] == 0),  # 19
        (filtered_df['country'] == "PL"),  # 20
        (filtered_df['country'] != "PL") # 21
    ]

    # List of result DataFrames
    result_dfs = [filtered_df.groupby('country')[['ilosc', 'wartosc_transakcji']].sum().reset_index()]

    # Apply conditions and perform groupby and sum operations
    for condition in conditions:
        result_df = filtered_df[condition].groupby('country')[['ilosc', 'wartosc_transakcji']].sum().reset_index()
        result_dfs.append(result_df)

    print(wb_sheets[1][0] + ' and ' + wb_sheets[2][0])
    s1 = wb_sheets[1][1]
    df1 = pd.DataFrame(s1.values)
    s2 = wb_sheets[2][1]
    df2 = pd.DataFrame(s1.values)

    for n in range(len(result_dfs) - 2):
        for i, country in enumerate(result_dfs[n]['country']):
            try:
                if country == 'PL':
                    country = 'W2'

                # Find column for country
                col1 = pd.Index(df1.iloc[27]).get_loc(country)
                col2 = pd.Index(df2.iloc[27]).get_loc(country)

                if AR2_4_row_1[n] in [30, 38, 44, 52, 61, 74, 76, 85]:
                    continue
                else:
                    # Insert value
                    s1[reference(col1, 'col') + reference(AR2_4_row_1[n], 'row')] = int(result_dfs[n]['ilosc'].iloc[i])
                    s2[reference(col2, 'col') + reference(AR2_4_row_2[n], 'row')] = to_float(
                        result_dfs[n]['wartosc_transakcji'].iloc[i])

            except KeyError:
                bug_table.append([f'BUG_4a.R.L_PLiW2', result_dfs[n][result_dfs[n]['country'] == country]])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {result_dfs[n][result_df[n]['country'] == country]}")

    # 6.ab.LiW
    print(wb_sheets[6][0])
    s6 = wb_sheets[6][1]
    df6 = pd.DataFrame(s6.values)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(19)

    col = 34

    # L6.1.4
    condition = df6.iloc[30:, 2] == 'L6.1.4'
    row = condition[condition].index[0]

    s6[reference(col, 'col') + reference(row, 'row')] = result_dfs[20]['ilosc'].iloc[0]

    # 6.1.4
    condition = df6.iloc[30:, 2] == 'W6.1.4'
    row = condition[condition].index[0]
    s6[reference(col, 'col') + reference(row, 'row')] = result_dfs[20]['wartosc_transakcji'].iloc[0]

    # L6.2.4
    condition = df6.iloc[30:, 2] == 'L6.2.4'
    row = condition[condition].index[0]

    s6[reference(col, 'col') + reference(row, 'row')] = result_dfs[21]['ilosc'].iloc[0]

    # 6.2.4
    condition = df6.iloc[30:, 2] == 'W6.2.4'
    row = condition[condition].index[0]
    s6[reference(col, 'col') + reference(row, 'row')] = result_dfs[21]['wartosc_transakcji'].iloc[0]

    if progress_callback_text:
        progress_callback_text(f'AR2 - Finished: 4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW.')

    # 5a.R.SF
    print(wb_sheets[5][0])
    if progress_callback_text:
        progress_callback_text(f'AR2 - Query: 5a.R.SF.')

    s5 = wb_sheets[5][1]
    df5 = pd.DataFrame(s5.values)

    temp_table = f"Query\\AR2\\NBP_Temp_3.sql"
    query = f"Query\\AR2\\NBP_Query_3.sql"

    dataframe_3 = load_or_query(18, '5a.R.SF__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(27)

    # 8.1.3.1
    condition = df5.iloc[30:, 2] == '8.1.3.1'
    row = condition[condition].index[0]
    s5[reference(5, 'col') + reference(row, 'row')] = dataframe_3[2]['wartosc'].iloc[0]

    # 8.1.3.2
    condition = df5.iloc[30:, 2] == '8.1.3.2'
    row = condition[condition].index[0]
    s5[reference(5, 'col') + reference(row, 'row')] = dataframe_3[3]['wartosc'].iloc[0]

    # 8.1.3.3
    condition = df5.iloc[30:, 2] == '8.1.3.3'
    row = condition[condition].index[0]
    s5[reference(5, 'col') + reference(row, 'row')] = dataframe_3[4]['wartosc'].iloc[0]

    if progress_callback_text:
        progress_callback_text(f'AR2 - Finished: 5a.R.SF.')

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(29)

    # 5a.R.LF_PLiW2 and 5a.R.WF_PLiW2
    print(wb_sheets[3][0] + ' and ' + wb_sheets[4][0])

    s3 = wb_sheets[3][1]
    df3 = pd.DataFrame(s3.values)
    s4 = wb_sheets[4][1]
    df4 = pd.DataFrame(s4.values)

    if progress_callback_text:
        progress_callback_text(f'AR2 - Fraud data: 5a.R.LF_PLiW2 and 5a.R.WF_PLiW2.')

    file_path = f"{TEMP}df_f.csv"

    if not os.path.exists(file_path):
        # Get the Visa
        data_visa = f_visa_make(user, passw)

        # Calculate and update progress to 30% when appropriate
        if progress_callback:
            progress_callback(35)

        df_visa = pd.DataFrame(data_visa[0])
        # Get the Mastercard
        data_mastercard = f_mastercard_make()

        # Calculate and update progress to 30% when appropriate
        if progress_callback:
            progress_callback(39)

        df_mastercard = pd.DataFrame(data_mastercard[0])
        df_complete = [df_visa, df_mastercard]
        df_fraud = pd.concat(df_complete)

        # Use extracted ARNs for query
        with open("./Query/AR2/NBP_Temp_2.sql", 'r',
                  encoding='utf-8') as sql:
            sql = sql.read()
            pattern = '--insert'
            insert_start = [match.start() for match in re.finditer(pattern, sql)]
            mastercard_insert = [insert_start[0], insert_start[0] + len(pattern)]
            visa_insert = [insert_start[1], insert_start[1] + len(pattern)]

            sql = sql[:visa_insert[0]] + data_visa[1] + sql[visa_insert[1]:]
            sql = sql[:mastercard_insert[0]] + data_mastercard[1] + sql[mastercard_insert[1]:]

        with open("./Query/AR2/NBP_Temp_2_filled.sql", 'w',
                  encoding='utf-8') as sql_w:
            sql_w.write(sql)

        # Calculate and update progress to 30% when appropriate
        if progress_callback:
            progress_callback(43)

        # Make pivot table
        df_fraud['country_aggr'] = df_fraud['country'].apply(lambda u: aggr_country(u))

        df_fraud.to_csv(f'{TEMP}df_fraud.csv')

        countries_list = [
            "BE",
            "IT",
            "AT",
            "BG",
            "HR",
            "CY",
            "CZ",
            "DK",
            "EE",
            "FI",
            "FR",
            "GR",
            "ES",
            "NL",
            "IE",
            "IS",
            "LT",
            "LI",
            "LU",
            "LV",
            "MT",
            "DE",
            "NO",
            "PT",
            "RO",
            "SK",
            "SI",
            "SE",
            "HU"
        ]

        country_mapping = {'PL': 'W2'}

        df_fraud['Category'] = df_fraud['country'].apply(
            lambda x: x if x in countries_list else country_mapping.get(x, 'G1'))

        query_arn = ''
        for number, arn in enumerate(df_fraud['ARN']):
            if number == df_fraud['ARN'].size - 1:
                query_arn += f"'{arn}'"
            else:
                query_arn += f"'{arn}',"

        sql_select = f"""
    SELECT
    aquirerReferenceNumber 'ARN',
    te_pos_entry_mode, 
    IIF(substring(te_pos_entry_mode, 1, 2) = '01', 'MOTO', 'inne') as czy_moto,
    CASE WHEN abs(tr_amount)<=10000 and substring(te_pos_entry_mode, 1, 2) in ('07','91') THEN 1
    ELSE 0 end as czy_niskokwotowa_zblizeniowa
    
    FROM [paytel_olap].[dbo].[v_trans] -- w przypadku starszych niż 90 dni v_trans
    join [paytel_olap].[dbo].v_trans_ext on (tr_tran_nr=te_tran_nr)	
    left join [paytel_olap].[dbo].v_visa_transaction  on (tranNr=tr_tran_nr) 
    left join [paytel_olap].[dbo].terminal on (_tr_tid=_t_tid)
    left join [paytel_olap].[dbo].shop on (ms_id=t_sid)
    left join [paytel_olap].[dbo].merchant on (tr_mid=m_mid)
    left join [paytel_olap].[dbo].visa_product_id on productId=vp_product_id
    WHERE aquirerReferenceNumber in ({query_arn})
    
    UNION
    
    SELECT
    aquirerReferenceNumber 'ARN',
    te_pos_entry_mode, 
    IIF(substring(te_pos_entry_mode, 1, 2) = '01', 'MOTO', 'inne') as czy_moto,
    CASE WHEN abs(tr_amount)<=10000 and substring(te_pos_entry_mode, 1, 2) in ('07','91') THEN 1
    ELSE 0 end as czy_niskokwotowa_zblizeniowa
    
    FROM [paytel_olap].[dbo].[v_trans] -- w przypadku starszych niż 90 dni v_trans
    join [paytel_olap].[dbo].v_trans_ext on (tr_tran_nr=te_tran_nr)	
    JOIN [paytel_olap].[dbo].if_transaction AS IT (NOLOCK) ON isnull(tr_prev_tran_nr, tr_tran_nr) = tranNr
    LEFT JOIN [paytel_olap].[dbo].mc_transaction AS MCT (NOLOCK)ON IT.postTranId = MCT.postTranId
    left join [paytel_olap].[dbo].terminal on (_tr_tid=_t_tid)
    left join [paytel_olap].[dbo].shop on (ms_id=t_sid)
    left join [paytel_olap].[dbo].merchant on (tr_mid=m_mid)
    
    WHERE aquirerReferenceNumber in ({query_arn})"""

        result = connect_single_query(sql_select)[0]
        df_fraud = pd.merge(df_fraud, result, on='ARN', how='inner')

        # Group by and calculate sum and count aggregations
        grouped_sum = \
            df_fraud.groupby(['Category', 'czy_SCA', 'FT description', 'Typ_karty', 'tr_sink_node', 'czy_moto',
                              'czy_niskokwotowa_zblizeniowa', 'pos_entry_mode'])['tr_amout'].sum().reset_index()
        grouped_count = \
            df_fraud.groupby(['Category', 'czy_SCA', 'FT description', 'Typ_karty', 'tr_sink_node', 'czy_moto',
                              'czy_niskokwotowa_zblizeniowa', 'pos_entry_mode'])[
                'tr_amout'].count().reset_index()

        # Rename columns for clarity
        grouped_sum.rename(columns={'tr_amout': 'Sum'}, inplace=True)
        grouped_count.rename(columns={'tr_amout': 'Count'}, inplace=True)

        # Merge the grouped dataframes on the specified columns
        merged_grouped_table = pd.merge(grouped_sum, grouped_count,
                                        on=['Category', 'czy_SCA', 'FT description', 'Typ_karty', 'tr_sink_node',
                                            'czy_moto', 'czy_niskokwotowa_zblizeniowa', 'pos_entry_mode'])

        # Set the desired index
        merged_grouped_table.set_index(['Category'], inplace=True)

        # 8 - Card-based payment transactions with card-based payment instruments issued by resident PSP (except cards
        # with an e-money function only) [received] Define columns to be summed
        # Define columns to be summed
        sum_columns = ['Sum', 'Count']

        df_total = merged_grouped_table.groupby(
            by=['Category', 'tr_sink_node', 'czy_moto', 'Typ_karty', 'czy_niskokwotowa_zblizeniowa', 'czy_SCA',
                'FT description'])[sum_columns].sum().reset_index()
    else:
        df_total = pd.read_csv(file_path)

    sheets = [wb_sheets[3][0], wb_sheets[4][0]]

    # 8.1.1.2 initiated via remote payment channel
    df_total_moto = df_total[df_total['czy_moto'] == 'MOTO']

    for sheet in sheets:
        for r in range(df_total_moto.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.1.2'
            row = condition[condition].index[0]

            country = df_total_moto.iloc[r][0]
            sum_v = df_total_moto.iloc[r][7]
            count_v = df_total_moto.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    df_total_moto = df_total[df_total['czy_moto'] == 'inne']

    for sheet in sheets:
        for r in range(df_total_moto.shape[0]):
            country = df_total_moto.iloc[r][0]
            sum_v = df_total_moto.iloc[r][7]
            count_v = df_total_moto.iloc[r][8]

            # 8.1.2.1.1.1 Initiated at a physical EFTPOS
            condition = df3.iloc[30:, 2] == '8.1.2.1.1.1'
            row = condition[condition].index[0]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # =================================== VISA

    # 8.1.2.1.2.1 Card-based payment instruments issued under PCS VISA
    df_total_tr_sink_node_visa = df_total[df_total['tr_sink_node'] == 'SN-Visa']

    # 8.1.2.1.2.1.1.1 with a debit card
    df_total_typ_karty_debit = df_total_tr_sink_node_visa[df_total_tr_sink_node_visa['Typ_karty'] == 'Debit']

    for sheet in sheets:
        for r in range(df_total_typ_karty_debit.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.1.1'
            row = condition[condition].index[0]

            country = df_total_typ_karty_debit.iloc[r][0]
            sum_v = df_total_typ_karty_debit.iloc[r][7]
            count_v = df_total_typ_karty_debit.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # 8.1.2.1.2.1.1.3 with a credit card
    df_total_typ_karty_credit = df_total_tr_sink_node_visa[df_total_tr_sink_node_visa['Typ_karty'] == 'Credit']

    for sheet in sheets:
        for r in range(df_total_typ_karty_credit.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.1.3'
            row = condition[condition].index[0]

            country = df_total_typ_karty_credit.iloc[r][0]
            sum_v = df_total_typ_karty_credit.iloc[r][7]
            count_v = df_total_typ_karty_credit.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # 8.1.2.1.2.1.2.1 SCA
    df_total_visa_sca = df_total_tr_sink_node_visa[df_total_tr_sink_node_visa['czy_SCA'] == 'SCA']

    for sheet in sheets:
        for r in range(df_total_visa_sca.shape[0]):
            # condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.1'
            # row = condition[condition].index[0]
            #
            country = df_total_visa_sca.iloc[r][0]
            sum_v = df_total_visa_sca.iloc[r][7]
            count_v = df_total_visa_sca.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if df_total_visa_sca.iloc[r]['FT description'] == 'Zgubienie lub kradzież karty':
                # 8.1.2.1.2.1.2.1.1.1 Lost or Stolen card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.1.1.1'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_sca.iloc[r]['FT description'] == 'Karta sfałszowana':
                # 8.1.2.1.2.1.2.1.1.3 Counterfeit card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.1.1.3'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_sca.iloc[r]['FT description'] == 'Nieodebrana karta':
                # 8.1.2.1.2.1.2.1.1.2 Card Not Received
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.1.1.2'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            else:
                # 8.1.2.1.2.1.2.1.1.4 Others
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.1.1.4'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

    # 8.1.2.1.2.1.2.2 non-SCA
    df_total_visa_non_sca = df_total_tr_sink_node_visa[df_total_tr_sink_node_visa['czy_SCA'] == 'non_SCA']

    for sheet in sheets:
        for r in range(df_total_visa_non_sca.shape[0]):
            # condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.2'
            # row = condition[condition].index[0]
            #
            country = df_total_visa_non_sca.iloc[r][0]
            sum_v = df_total_visa_non_sca.iloc[r][7]
            count_v = df_total_visa_non_sca.iloc[r][8]
            #
            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if df_total_visa_non_sca.iloc[r]['FT description'] == 'Zgubienie lub kradzież karty':
                # 8.1.2.1.2.1.2.2.1.1 Lost or Stolen card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.2.1.1'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_non_sca.iloc[r]['FT description'] == 'Karta sfałszowana':
                # 8.1.2.1.2.1.2.2.1.3 Counterfeit card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.2.1.3'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = to_float(df4[col].iloc[row]) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_non_sca.iloc[r]['FT description'] == 'Nieodebrana karta':
                # 8.1.2.1.2.1.2.2.1.2 Card Not Received
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.2.1.2'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)
            else:
                # 8.1.2.1.2.1.2.2.1.4 Others
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.1.2.2.1.4'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

    # =================================== MASTERCARD

    # 8.1.2.1.2.2 Card-based payment instruments issued under PCS MASTERCARD
    df_total_tr_sink_node_masterc = df_total[df_total['tr_sink_node'] == 'SN-MasterC']

    # 8.1.2.1.2.2.1.1 with a debit card
    df_total_typ_karty_debit = df_total_tr_sink_node_masterc[df_total_tr_sink_node_masterc['Typ_karty'] == 'Debit']

    for sheet in sheets:
        for r in range(df_total_typ_karty_debit.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.1.1'
            row = condition[condition].index[0]

            country = df_total_typ_karty_debit.iloc[r][0]
            sum_v = df_total_typ_karty_debit.iloc[r][7]
            count_v = df_total_typ_karty_debit.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # 8.1.2.1.2.2.1.3 with a credit card
    df_total_typ_karty_credit = df_total_tr_sink_node_masterc[
        df_total_tr_sink_node_masterc['Typ_karty'] == 'Credit']

    for sheet in sheets:
        for r in range(df_total_typ_karty_credit.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.1.3'
            row = condition[condition].index[0]

            country = df_total_typ_karty_credit.iloc[r][0]
            sum_v = df_total_typ_karty_credit.iloc[r][7]
            count_v = df_total_typ_karty_credit.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # 8.1.2.1.2.2.2.1 SCA
    df_total_masterc_sca = df_total_tr_sink_node_masterc[df_total_tr_sink_node_masterc['czy_SCA'] == 'SCA']

    for sheet in sheets:
        for r in range(df_total_masterc_sca.shape[0]):

            country = df_total_masterc_sca.iloc[r][0]
            sum_v = df_total_masterc_sca.iloc[r][7]
            count_v = df_total_masterc_sca.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if df_total_visa_sca.iloc[r]['FT description'] == 'Zgubienie lub kradzież karty':
                # 8.1.2.1.2.2.2.1.1.1 Lost or Stolen card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.1.1.1'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = to_float(df4[col].iloc[row]) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_sca.iloc[r]['FT description'] == 'Karta sfałszowana':
                # 8.1.2.1.2.2.2.1.1.3 Counterfeit card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.1.1.3'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = to_float(df4[col].iloc[row]) + round(
                        to_float(sum_v), 2)

            elif df_total_visa_sca.iloc[r]['FT description'] == 'Nieodebrana karta':
                # 8.1.2.1.2.2.2.1.1.2 Card Not Received
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.1.1.2'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)
            else:
                # 8.1.2.1.2.1.2.1.1.4 Others
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.1.1.4'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

    # 8.1.2.1.2.2.2.2 non-SCA
    df_total_masterc_non_sca = df_total_tr_sink_node_masterc[df_total_tr_sink_node_masterc['czy_SCA'] == 'non_SCA']

    for sheet in sheets:
        for r in range(df_total_masterc_non_sca.shape[0]):

            country = df_total_masterc_non_sca.iloc[r][0]
            sum_v = df_total_masterc_non_sca.iloc[r][7]
            count_v = df_total_masterc_non_sca.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if df_total_masterc_non_sca.iloc[r]['FT description'] == 'Zgubienie lub kradzież karty':
                # 8.1.2.1.2.2.2.2.1.1 Lost or Stolen card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.2.1.1'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            elif df_total_masterc_non_sca.iloc[r]['FT description'] == 'Karta sfałszowana':
                # 8.1.2.1.2.2.2.2.1.3 Counterfeit card
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.2.1.3'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

            elif df_total_masterc_non_sca.iloc[r]['FT description'] == 'Nieodebrana karta':
                # 8.1.2.1.2.2.2.2.1.2 Card Not Received
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.2.1.2'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)
            else:
                # 8.1.2.1.2.2.2.2.1.4 Others
                condition = df3.iloc[30:, 2] == '8.1.2.1.2.2.2.2.1.4'
                row = condition[condition].index[0]

                if sheet == '5a.R.LF_PLiW2':
                    s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                        to_float(count_v), 2)
                else:
                    s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                        to_float(sum_v), 2)

    # 8.1.2.1.3.3 Contactless low value
    df_total_contactless = df_total[df_total['czy_niskokwotowa_zblizeniowa'] == 1]
    df_total_contactless = df_total_contactless[df_total_contactless['czy_SCA'] == 'non_SCA']

    for sheet in sheets:
        for r in range(df_total_contactless.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.3.3'
            row = condition[condition].index[0]

            country = df_total_contactless.iloc[r][0]
            sum_v = df_total_contactless.iloc[r][7]
            count_v = df_total_contactless.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # 8.1.2.1.3.5 Other
    df_total_contactless = df_total[df_total['czy_niskokwotowa_zblizeniowa'] == 0]
    df_total_contactless = df_total_contactless[df_total_contactless['czy_SCA'] == 'non_SCA']

    for sheet in sheets:
        for r in range(df_total_contactless.shape[0]):
            condition = df3.iloc[30:, 2] == '8.1.2.1.3.5'
            row = condition[condition].index[0]

            country = df_total_contactless.iloc[r][0]
            sum_v = df_total_contactless.iloc[r][7]
            count_v = df_total_contactless.iloc[r][8]

            try:
                col = pd.Index(df3.iloc[27]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', country])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {country}")

            if sheet == '5a.R.LF_PLiW2':
                s3[reference(col, 'col') + reference(row, 'row')] = round(to_float(df3[col].iloc[row]), 2) + round(
                    to_float(count_v), 2)
            else:
                s4[reference(col, 'col') + reference(row, 'row')] = round(to_float(df4[col].iloc[row]), 2) + round(
                    to_float(sum_v), 2)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(47)
    if progress_callback_text:
        progress_callback_text(f'AR2 - Finished: 5a.R.LF_PLiW2 and 5a.R.WF_PLiW2.')

    print('\nChecking the quarter: ' + str(check_quarter()[3]))

    # 9.R.L.MCC and 9.R.W.MCC
    print(wb_sheets[7][0] + ' and ' + wb_sheets[8][0])

    temp_table = f"Query\\AR2\\NBP_Temp_4.sql"
    query = f"Query\\AR2\\NBP_Query_4.sql"

    if progress_callback_text:
        progress_callback_text(f'AR2 - Query: 9.R.L.MCC and 9.R.W.MCC.')
    # dataframe_4 = load_or_query(4, '9.R.L.MCC_9.R.W.MCC__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(49)

    s7 = wb_sheets[7][1]
    df7 = pd.DataFrame(s7.values)
    s8 = wb_sheets[8][1]
    df8 = pd.DataFrame(s8.values)

    code = [9, '9.1', '9.1.1']

    def change_country(val):
        if val in ["PL", "PR", "QZ"]:
            change = {
                "PL": "W2",
                "PR": "US",
                "QZ": "D09"
            }
            return change[val]
        else:
            return val

    mcc_index = df7[2][35:370].astype(str)

    def change_mcc(val):
        if str(val) in ['742', '743', '744', '763', '780']:
            mcc_change = {'742': '0742',
                          '743': '0743',
                          '744': '0744',
                          '763': '0763',
                          '780': '0780'}
            return mcc_change[str(val)]

        elif mcc_index[mcc_index == str(val)].index.size == 0:
            return "R999"

        else:
            return str(val)

    df_for_mcc['mcc'] = df_for_mcc['mcc'].apply(change_mcc)

    df_mcc_moto = df_for_mcc.groupby(['country', 'mcc', 'czy_moto'])[
        ['ilosc', 'wartosc_transakcji']].sum().reset_index()

    df_mcc = df_for_mcc.groupby(['country', 'mcc'])[['ilosc', 'wartosc_transakcji']].sum().reset_index()

    df_mcc_moto['country'] = df_mcc_moto['country'].apply(change_country)
    df_mcc['country'] = df_mcc['country'].apply(change_country)

    pd.DataFrame(df_mcc_moto).to_excel(f'{TEMP}df_mcc_moto.xlsx')
    pd.DataFrame(df_mcc).to_excel(f'{TEMP}df_mcc.xlsx')

    for country in df_mcc['country']:

        # 9
        col = pd.Index(df7.iloc[27]).get_loc(country)
        condition = df7[1][30:370] == code[0]
        row = condition[condition].index[0]

        addr = reference(col, 'col') + reference(row, 'row')

        result_1 = df_mcc[df_mcc['country'] == country]['ilosc'].sum()
        result_2 = df_mcc[df_mcc['country'] == country]['wartosc_transakcji'].sum()
        print('result 1 and 2: ', result_1, result_2)
        s7[addr].value = result_1
        s8[addr].value = result_2

        # 9.1
        condition = df7[1][30:370] == code[1]
        row = condition[condition].index[0]

        addr = reference(col, 'col') + reference(row, 'row')

        result_1 = df_mcc_moto[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')]['ilosc'].sum()
        result_2 = df_mcc_moto[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')][
            'wartosc_transakcji'].sum()
        print('result 1 and 2: ', result_1, result_2)
        s7[addr].value = result_1
        s8[addr].value = result_2

        # 9.1.1
        condition = df7[1][30:370] == code[2]
        row = condition[condition].index[0]

        addr = reference(col, 'col') + reference(row, 'row')

        result_1 = df_mcc_moto[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')]['ilosc'].sum()
        result_2 = df_mcc_moto[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')][
            'wartosc_transakcji'].sum()
        print('result 1 and 2: ', result_1, result_2)
        s7[addr].value = result_1
        s8[addr].value = result_2

        df7[2] = df7[2].astype(str)

        mccs = df_mcc[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')]['mcc'].astype(str)

        for m, mcc in enumerate(mccs):
            print(mcc, type(mcc))
            # Convert df_nbp_2[1] column to string
            mcc_value_w = df_mcc[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')]['wartosc_transakcji'].iloc[m]
            mcc_value_i = df_mcc[(df_mcc_moto['country'] == country) & (df_mcc_moto['czy_moto'] == 'inne')]['ilosc'].iloc[m]

            try:
                col = pd.Index(df7.iloc[27]).get_loc(country)

                if mcc_index[mcc_index == mcc].index.size == 0:
                    print(mcc, ': R999', 'mcc_am :', mcc_value_i)
                    ind = mcc_index[mcc_index == "R999"].index[0]

                    cell_reference = reference(col, 'col') + reference(ind, 'row')

                    if s7[cell_reference].value is not None:
                        print(s7[cell_reference].value)
                        s7[cell_reference].value += int(mcc_value_i)
                    else:
                        s7[cell_reference].value = int(mcc_value_i)

                    if s8[cell_reference].value is not None:
                        s8[cell_reference].value += float(mcc_value_w)
                    else:
                        s8[cell_reference].value = float(mcc_value_w)
                else:
                    print(mcc)
                    ind = mcc_index[mcc_index == mcc].index[0]

                    cell_reference = reference(col, 'col') + reference(ind, 'row')
                    s7[cell_reference] = mcc_value_i
                    s8[cell_reference] = mcc_value_w

            except (KeyError, IndexError) as e:

                if country not in pd.Index(df7.iloc[27]):
                    bug_table.append([f'BUG_{wb_sheets[7][0]}', country])
                    print(
                        f"{e} / {wb_sheets[7][0]}!!: Value was not added to the report (there is no such a country code in excel) - {country}")

                elif mcc_index[mcc_index == mcc].index.size == 0:
                    ind = mcc_index[mcc_index == "R999"].index[0]
                    # s7[reference(col, 'col') + reference(ind, 'row')] = mcc_value_i
                    # s8[reference(col, 'col') + reference(ind, 'row')] = mcc_value_w

                    print(
                        f'{e} / {wb_sheets[7][0]}!!:For country: {country} and mcc: {mcc} added it to not assigned mccs R999.')

    if progress_callback_text:
        progress_callback_text(f'AR2 - Finished: 9.R.L.MCC and 9.R.W.MCC.')

    if "df_fraud" in locals():
        df_fraud.to_csv(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/df_f.csv')


def prepare_data_ar1(user, passw, df_f, name, surname, phone, email, progress_callback=None,
                     progress_callback_text=None):
    wb_data = create_ar1_excel()

    # Create new AR2 excel file
    sheets = wb_data[0]
    wb = wb_data[1]

    # ST.01
    print(sheets[0][0])
    s1 = sheets[0][1]

    if progress_callback_text:
        progress_callback_text(f'AR1 - Query: ST.01.')

    temp_table = f"Query\\AR1\\NBP_Temp_1.sql"
    query = f"Query\\AR1\\NBP_Query_1.sql"

    dataframe_1 = load_or_query(2, 'ST.01.', temp_table, query)

    if progress_callback_text:
        progress_callback_text(f'AR1 - Finished: ST.01.')
    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(55)

    # Get data from "Tvid_nev_lost.xlsx'
    path = f'//prdfil/Business/DPiUS/Zespol Przetwarzania/Raporty kwartalne/{check_quarter()[1]}Q{check_quarter()[3]}/Tvid_nev_lost.xlsx'
    read_remote_file(path, user, passw)
    dataframe_0 = pd.read_excel(path, header=3)

    print('\nData z pliku Tvid_nev_lost.xlsx: ', dataframe_0['tr_date'][0])

    # Data to be filled
    to_change_values = [
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_mid'][0],  # mid all
        dataframe_1[1]['MID all cashback'][0],

        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_1[1]['SID all cashback'][0],

        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_1[1]['TVID all cashback'][0],
        dataframe_0['active_tvid'][0],  # tvid all

        dataframe_1[0]['ilosc_softPOS'][0],


    ]
    to_change_rows = [9, 10, 11, 14, 15, 16, 18, 19, 20, 21, 22, 24]
    to_change_column = 9

    # Changes in dataframe from spreadsheet - df_nbp_1
    for v in range(len(to_change_values)):
        s1[reference(to_change_column, 'col') + reference(to_change_rows[v], 'row')] = to_change_values[v]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(57)

    # ST.03
    print(sheets[1][0])
    s3 = sheets[1][1]

    to_change_values = [
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0]  # tvid all

    ]
    to_change_rows = [9, 10, 14, 15, 18, 19]
    to_change_column = 13

    # Just using values filled to sheet ST.03
    for v in range(len(to_change_values)):
        s3[reference(to_change_column, 'col') + reference(to_change_rows[v], 'row')] = to_change_values[v]

    if progress_callback_text:
        progress_callback_text(f'AR1 - Finished: ST.03.')

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(64)

    # ST.05
    print(sheets[2][0])
    s5 = sheets[2][1]

    temp_table = f"Query\\AR1\\NBP_Temp_2.sql"
    query = f"Query\\AR1\\NBP_Query_2.sql"

    if progress_callback_text:
        progress_callback_text(f'AR1 - Query: ST.05.')

    dataframe_2 = load_or_query(2, 'ST.05.', temp_table, query)

    temp_table = f"Query\\AR2\\NBP_Temp_1.sql"
    query = f"Query\\AR2\\NBP_Query_1.sql"

    dataframe_1 = load_or_query(1, '4.a.R.L_PLiW2_4a.R.W_PLiW2_6.ab.LiW__', temp_table, query)

    # Remove spaces from country code
    dataframe_1[0]['country'] = dataframe_1[0]['country'].str.replace(' ', '')

    # Function to replace values based on condition
    def replace_country(value):
        if len(value) > 2:
            return COUNTRY_DICT.get(value, value)
        else:
            return value

    # Apply the function to the first column
    dataframe_1[0]['country'] = dataframe_1[0]['country'].apply(replace_country)

    df_clean = dataframe_1[0].copy()

    # Columns:
    # (0) karta / (1) country / (2) mcc / (3) Typ_karty / (4) te_pos_entry_mode / (5) czy_moto /
    # (6) czy_niskokwotowa_zblizeniowa / (7) czy_SCA / (8) ilosc / (9) wartosc_transakcji / (10) te_tran_type
    # (11) Wartość wypłat Cash Back / (12) category / (13) tr_rsp_code / (14) tr_app_id

    def polska_inne(value):
        if value == 'PL':
            return 'POLSKA'
        else:
            return 'INNE KRAJE'

    def business_individual(value):
        if value != 'BUSINESS':
            return 'Individual'
        else:
            return 'BUSINESS'

    # Filter the DataFrame based on the list of countries
    dataframe_1[0]['country'] = dataframe_1[0]['country'].apply(polska_inne)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(68)

    dataframe_3 = df_clean

    def change_country(val):
        if val in {"PL", "PR", "QZ"}:
            change = {
                "PL": "W2",
                "PR": "US",
                "QZ": "D09"
            }
            return change[val]
        else:
            return val

    dataframe_3['country'] = dataframe_3['country'].apply(change_country)

    dataframe_3['ilosc_internet'] = 0
    dataframe_3['wartosc_internet'] = 0
    dataframe_3['ilosc_cashback'] = np.where(dataframe_3['Wartość wypłat Cash Back'] != 0, 1, 0)

    dataframe_3 = dataframe_3.groupby(['country', 'category', 'te_pos_entry_mode']).agg({
        'Wartość wypłat Cash Back': 'sum',
        'ilosc_cashback': 'sum',
        'ilosc_internet': 'sum',
        'wartosc_internet': 'sum',
        'ilosc': 'sum',
        'wartosc_transakcji': 'sum'
    }).reset_index()

    geo6 = pd.read_excel("./Example/NBP_GEO6.xlsx", header=3)

    dataframe_3 = pd.merge(dataframe_3, geo6, on='country', how='left')
    dataframe_3.to_excel('ST.05_data.xlsx')
    # Data to be filled

    def prepare_values_data(d, c, result_df_1):
        if d == 0:
            # 1 / "Liczba transakcji CashBack" / "Wartość wypłat Cash Back"
            # row 12 / 16

            # Print columns of result_df_1
            print(result_df_1.columns)

            conditions = [
                (result_df_1['category'] == c) & (result_df_1['country'] == 'W2'),
                (result_df_1['category'] == c) & (result_df_1['country'] != 'W2'),
                (result_df_1['category'] == c) & (result_df_1['country'] == 'W2'),
                (result_df_1['category'] == c) & (result_df_1['country'] != 'W2')
            ]

            return [
                result_df_1.loc[conditions[0], 'ilosc_cashback'].sum(),
                result_df_1.loc[conditions[1], 'ilosc_cashback'].sum(),
                result_df_1.loc[conditions[2], 'Wartość wypłat Cash Back'].sum(),
                result_df_1.loc[conditions[3], 'Wartość wypłat Cash Back'].sum(),
                result_df_1.loc[conditions[0], 'ilosc_cashback'].sum(),
                result_df_1.loc[conditions[1], 'ilosc_cashback'].sum(),
                result_df_1.loc[conditions[2], 'Wartość wypłat Cash Back'].sum(),
                result_df_1.loc[conditions[3], 'Wartość wypłat Cash Back'].sum()
            ]

        elif d == 1:
            # 2 / "Liczba transakcji BLIK" / "Kwota transakcji BLIK"
            # row 20
            ilosc = 'Liczba transakcji BLIK'
            wartosc = 'Kwota transakcji BLIK'

            return [
                dataframe_2[0][ilosc][0],
                0,
                dataframe_2[0][wartosc][0],
                0,
                dataframe_2[0][ilosc][0],
                0,
                dataframe_2[0][wartosc][0],
                0
            ]
        elif d == 2:
            # 3
            # row 10 / 14
            df_bi = result_df_1
            df_bi['category'] = df_bi['category'].apply(business_individual)

            result_df_2 = df_bi

            conditions = [
                (result_df_2['category'] == c) & (result_df_2['country'] == 'W2'),
                (result_df_2['category'] == c) & (result_df_2['country'] != 'W2'),
                (result_df_2['category'] == c) & (result_df_2['country'] == 'W2'),
                (result_df_2['category'] == c) & (result_df_2['country'] != 'W2')
            ]

            return [
                result_df_2.loc[conditions[0], 'ilosc'].sum(),
                result_df_2.loc[conditions[1], 'ilosc'].sum(),
                result_df_2.loc[conditions[2], 'wartosc_transakcji'].sum(),
                result_df_2.loc[conditions[3], 'wartosc_transakcji'].sum(),
                result_df_2.loc[conditions[0], 'ilosc'].sum(),
                result_df_2.loc[conditions[1], 'ilosc'].sum(),
                result_df_2.loc[conditions[2], 'wartosc_transakcji'].sum(),
                result_df_2.loc[conditions[3], 'wartosc_transakcji'].sum()

            ]

        elif d == 3:
            # 4
            # row 11 / 15
            result_df_3 = result_df_1[result_df_1['te_pos_entry_mode'] == 'CTLS']

            conditions = [
                (result_df_3['category'] == c) & (result_df_3['country'] == 'W2'),
                (result_df_3['category'] == c) & (result_df_3['country'] != 'W2'),
                (result_df_3['category'] == c) & (result_df_3['country'] == 'W2'),
                (result_df_3['category'] == c) & (result_df_3['country'] != 'W2')
            ]

            return [
                result_df_3.loc[conditions[0], 'ilosc'].sum(),
                result_df_3.loc[conditions[1], 'ilosc'].sum(),
                result_df_3.loc[conditions[2], 'wartosc_transakcji'].sum(),
                result_df_3.loc[conditions[3], 'wartosc_transakcji'].sum(),
                result_df_3.loc[conditions[0], 'ilosc'].sum(),
                result_df_3.loc[conditions[1], 'ilosc'].sum(),
                result_df_3.loc[conditions[2], 'wartosc_transakcji'].sum(),
                result_df_3.loc[conditions[3], 'wartosc_transakcji'].sum()

            ]
        elif d == 4:
            # 5
            # c == 'Individual' 26
            # c == 'BUSINESS' 28
            d = 1
            ilosc = 'ilosc'
            wartosc = 'wartosc'
            return [
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),

                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum()
            ]

    to_change_columns = range(13, 22)
    to_change_rows_1 = [12, 20, 10, 11, 26]
    to_change_rows_2 = [16, 20, 14, 15, 28]

    # Changes in dataframe from spreadsheet - df_nbp_1
    for e in range(4):
        for c in ['Individual', 'BUSINESS']:
            to_change_values = prepare_values_data(e, c, dataframe_3)
            for v in range(len(to_change_values)):
                if c == 'Individual':
                    s5[reference(to_change_columns[v], 'col') + reference(to_change_rows_1[e], 'row')] = \
                        to_change_values[v]
                else:
                    s5[reference(to_change_columns[v], 'col') + reference(to_change_rows_2[e], 'row')] = \
                        to_change_values[v]

    if progress_callback_text:
        progress_callback_text(f'AR1 - Finished: ST.05.')

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(75)

    # ST.06
    print(sheets[3][0])
    s6 = sheets[3][1]
    df6 = pd.DataFrame(s6.values)

    temp_table = f"Query\\AR1\\NBP_Temp_3.sql"
    query = f"Query\\AR1\\NBP_Query_3.sql"

    if progress_callback_text:
        progress_callback_text(f'AR1 - Query: ST.06.')

    # dataframe_3 = load_or_query(2, 'ST.06.', temp_table, query)
    # print('\nDate: ' + str(dataframe_3[0]))

    dataframe_3 = dataframe_3[dataframe_3['country'] != 'W2'].groupby(['country']).agg({
        'ilosc': 'sum',
        'ilosc_internet': 'sum',
        'ilosc_cashback': 'sum',
        'wartosc_transakcji': 'sum',
        'wartosc_internet': 'sum',
        'Wartość wypłat Cash Back': 'sum'
    }).reset_index()

    dataframe_3.to_excel('ST.06_data.xlsx')

    print(dataframe_3.columns)

    column_amount = [13, 14, 15]
    column_value = [16, 17, 18]

    content_amount = ['ilosc', 'ilosc_internet', 'ilosc_cashback']
    content_value = ['wartosc_transakcji', 'wartosc_internet', 'Wartość wypłat Cash Back']

    # devices that accept payment cards / Internet / cash back
    for k, country in enumerate(dataframe_3['country']):
        print('1: ', geo6['country'].isin([country]).any())

        # IF country not in Database
        if country == 'uwaga - coś nowego':
            print(
                f"!!: In the report fraud transactions were not added (probably due to NULL value) - {dataframe_3[dataframe_3['country'] == 'uwaga - coś nowego']}")
            bug_table.append([f'ST.06', dataframe_3[dataframe_3['country'] == 'uwaga - coś nowego']])
        # IF country in Database AND within predefined countries in excel.

        elif country == 'PL':
            continue

        elif df6[0][9:40].isin([country]).any():
            row = df6[0][9:40][df6[0][9:40] == country].index[0]
            for i in range(3):
                df6.iat[row, column_amount[i]] = dataframe_3[dataframe_3['country'] == country][
                    content_amount[i]].values[0]
                df6.iat[row, column_value[i]] = dataframe_3[dataframe_3['country'] == country][
                    content_value[i]].values[0]
        # ELSE rest of countries - not predefined but within database
        else:
            # CASE when Nambia - in database its stored as NA which cause Null value for DataFrame
            if country == 'NA':
                c_name = 'Namibia'
                new_row_data = ['NA', c_name, c_name]

                for i in range(3):
                    new_row_data.append(
                        dataframe_3.iloc[k][content_amount[i]])
                for i in range(3):
                    new_row_data.append(
                        dataframe_3.iloc[k][content_value[i]])

                dataframe_3.iloc[k]['Name'] = 'Namibia'
                dataframe_3.iloc[k]['name_PL'] = 'Namibia'

            # ELSE - countries in database and not predefined in excel.
            else:
                print(country)
                if not geo6['country'].isin([country]).any():
                    country = 'D09'

                # Country name
                country_name = geo6[geo6['country'] == country]['name_PL'].values[0]
                # Country name - english version
                country_name_english = geo6[geo6['country'] == country]['Name'].values[0]
                # Add new row data to DataFrame
                new_row_data = [country, country_name, country_name_english]

                for i in range(3):
                    new_row_data.append(dataframe_3[dataframe_3['country'] == country][content_amount[i]].values[0])
                for i in range(3):
                    new_row_data.append(dataframe_3[dataframe_3['country'] == country][content_value[i]].values[0])

            columns = [0, 11, 12, 13, 14, 15, 16, 17, 18]
            first_null_index = df6[9:][0].index[df6[9:][0].isnull()].tolist()[0]

            for n, c in enumerate(columns):
                df6.iat[first_null_index, c] = new_row_data[n]

    last_index = df6[9:][0].index[df6[9:][0].isnull()].tolist()[0]
    rows_1 = [*range(9, 38)]
    rows_2 = [*range(39, last_index)]
    df6.to_excel('df6.xlsx')

    for rows in [rows_1, rows_2]:
        for row in rows:
            if row < 39:
                columns_excel = [13, 14, 15, 16, 17, 18]
            else:
                columns_excel = [0, 11, 12, 13, 14, 15, 16, 17, 18]
            for c in columns_excel:
                s6[reference(c, 'col') + reference(row, 'row')] = df6.iat[row, c]
                s6[reference(c, 'col') + reference(row, 'row')] = df6.iat[row, c]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(86)
    if progress_callback_text:
        progress_callback_text(f'AR1 - Finished: ST.06.')
    # ST.02

    # Right now we do not fill this sheet.

    # ST.07
    print(sheets[4][0])
    s7 = sheets[4][1]

    temp_table = f"Query\\AR1\\NBP_Temp_4.sql"
    query = f"Query\\AR1\\NBP_Query_4.sql"

    if progress_callback_text:
        progress_callback_text(f'AR1 - Query: ST.07.')

    dataframe_4 = load_or_query(4, 'ST.07.', temp_table, query)
    # Calculate and update progress to 30% when appropriate

    if progress_callback:
        progress_callback(98)

    dataframe_4[1]['kwota'] = dataframe_4[1]['kwota'].astype('float')
    s7[reference(13, 'col') + reference(14, 'row')] = dataframe_4[1][dataframe_4[1]['kraj'] == 'PL']['ilosc'].iloc[0]
    s7[reference(14, 'col') + reference(14, 'row')] = dataframe_4[1][dataframe_4[1]['kraj'] == 'other']['ilosc'].iloc[0]
    s7[reference(15, 'col') + reference(14, 'row')] = dataframe_4[1][dataframe_4[1]['kraj'] == 'PL']['kwota'].iloc[0]
    s7[reference(16, 'col') + reference(14, 'row')] = dataframe_4[1][dataframe_4[1]['kraj'] == 'other']['kwota'].iloc[0]
    s7[reference(17, 'col') + reference(14, 'row')] = dataframe_4[3]['kwota'].iloc[0]
    s7[reference(18, 'col') + reference(14, 'row')] = dataframe_4[2]['kwota'].iloc[0]

    dff = pd.pivot_table(index='pos_entry_mode', columns='country_aggr',
                         data=df_f[df_f['quarter'] == check_quarter()[3]],
                         aggfunc={'tr_amout': 'sum', 'country_aggr': 'count'}, fill_value=0)

    s7[reference(13, 'col') + reference(11, 'row')] = dff['country_aggr'].sum()['PL']
    s7[reference(14, 'col') + reference(11, 'row')] = dff['country_aggr'].sum()['NPL']
    s7[reference(15, 'col') + reference(11, 'row')] = dff['tr_amout'].sum()['PL']
    s7[reference(16, 'col') + reference(11, 'row')] = dff['tr_amout'].sum()['NPL']
    s7[reference(17, 'col') + reference(11, 'row')] = 0
    s7[reference(18, 'col') + reference(11, 'row')] = 0

    s7[reference(13, 'col') + reference(12, 'row')] = dff['country_aggr'].loc['CTLS']['PL']
    s7[reference(14, 'col') + reference(12, 'row')] = dff['country_aggr'].loc['CTLS']['NPL']
    s7[reference(15, 'col') + reference(12, 'row')] = dff['tr_amout'].loc['CTLS']['PL']
    s7[reference(16, 'col') + reference(12, 'row')] = dff['tr_amout'].loc['CTLS']['NPL']
    s7[reference(17, 'col') + reference(12, 'row')] = 0
    s7[reference(18, 'col') + reference(12, 'row')] = 0

    for n in range(13, 19):
        s7[reference(n, 'col') + reference(13, 'row')] = to_float(s7[reference(n, 'col') + reference(14, 'row')].value)

    for n in range(13, 19):
        s7[reference(n, 'col') + reference(10, 'row')] = to_float(
            s7[reference(n, 'col') + reference(11, 'row')].value) + to_float(
            s7[reference(n, 'col') + reference(13, 'row')].value)

    for n in range(13, 19):
        s7[reference(n, 'col') + reference(9, 'row')] = to_float(s7[reference(n, 'col') + reference(10, 'row')].value)

    if progress_callback_text:
        progress_callback_text(f'AR1 - Finished: ST.07.')

    # p-dane
    print(sheets[5][0])
    sp_dane = sheets[5][1]

    author_data = [
        name,
        surname,
        phone,
        email
    ]

    for x in range(13, 20):
        i = 0
        for y in range(9, 13):
            sp_dane[reference(x, 'col') + reference(y, 'row')] = author_data[i]
            i += 1

    # Save everything to new excel file - path to new file stored in wb_data[2]
    wb.save(wb_data[2])

    print('END')


def bug_report():
    wb_sheet_names = []
    if not bug_table:
        print('Full success - no bugs!')
    else:
        wb_bug = pd.ExcelWriter(f'{TEMP}bug_table.xlsx', engine='openpyxl')

        for i, bug in enumerate(bug_table):
            if bug[0] + '_0' not in wb_sheet_names:
                sheet_name = bug[0] + '_0'
                wb_sheet_names.append(sheet_name)
                bug[1].to_excel(wb_bug, sheet_name=sheet_name, index=False)
                print(f'Added bug table to sheet: {sheet_name}')
            else:
                no = []
                for sheet_name in wb_sheet_names:
                    if bug[0] + '_' in sheet_name:
                        try:
                            no.append(int(sheet_name[len(bug[0] + '_'):]))
                        except IndexError:
                            print('Error in writing the bug table.')

                max_no = max(no) + 1
                sheet_name = bug[0] + '_' + str(max_no)
                wb_sheet_names.append(sheet_name)
                bug[1].to_excel(wb_bug, sheet_name=sheet_name, index=False)
                print(f'Added bug table to sheet: {sheet_name}')

        # Add all the bug tables to a single sheet
        combined_sheet_name = 'Combined Sheet'
        all_bugs = pd.DataFrame(columns=['Sheet Name', 'Data', 'Sum'])

        for sheet_name in wb_sheet_names:
            data = bug_table[wb_sheet_names.index(sheet_name)][1]
            if 'wartosc_transakcji' in data.columns and 'wartosc_internet' in data.columns and 'wartosc_wyplat_CashBack' in data.columns:
                data_sum = data['wartosc_transakcji'] + data['wartosc_internet'] + data['wartosc_wyplat_CashBack']
                data_sum = data_sum.item() if len(data_sum) == 1 else None  # Extract the float value from Series
            else:
                data_sum = None

            row_data = {'Sheet Name': sheet_name, 'Data': data, 'Sum': data_sum}
            all_bugs = pd.concat([all_bugs, pd.DataFrame([row_data])], ignore_index=True)

        # Reorder the sheets
        writer = pd.ExcelWriter(f'{TEMP}bug_table.xlsx', engine='openpyxl')
        writer.book = wb_bug.book
        all_bugs['Sum'] = all_bugs['Sum'].astype(float)  # Convert the 'Sum' column to float
        all_bugs.to_excel(writer, sheet_name=combined_sheet_name, index=False)
        writer.save()

        print(f'Added all bug tables to the combined sheet: {combined_sheet_name}')

        wb_bug.close()
        print('Saved bug tables to bug_table.xlsx')


def create_ar2_excel():
    new_file_path = f"{PATH}/Filled/{check_quarter()[1]}_Q{check_quarter()[3]}___BSP_AR2_v.4.01.xlsx"
    shutil.copyfile(f"{PATH}TEMPLATE___BSP_AR2_v.4.01.xlsx", new_file_path)
    nbp_wb_2 = openpyxl.load_workbook(new_file_path, read_only=False, keep_links=False, data_only=False)
    sheets = []

    for sheet in EXCEL_READ_AR2:
        sheets.append([sheet, nbp_wb_2[sheet]])

    return sheets, nbp_wb_2, new_file_path


def create_ar1_excel():
    new_file_path = f"{PATH}/Filled/{check_quarter()[1]}_Q{check_quarter()[3]}___BSP_AR1_ST.w.8.7.5.xlsx"
    shutil.copyfile(f"{PATH}TEMPLATE___BSP_AR1_ST.w.8.7.5.xlsx", new_file_path)
    nbp_wb_1 = openpyxl.load_workbook(new_file_path, read_only=False, keep_links=False, data_only=False)
    sheets = []

    for sheet in EXCEL_READ_AR1:
        sheets.append([sheet, nbp_wb_1[sheet]])

    return sheets, nbp_wb_1, new_file_path


def start_automation(d1, d2, d3, d4, d_pass, progress_callback=None, progress_callback_text=None):
    # Open the log file in append mode
    log_file = open(to_log(), "a")

    # Create the logger object
    logger = Logger(log_file)

    # Create an instance of the Logger class
    sys_logger = Logger(sys.stdout)

    # Assign the logger as the new sys.stdout
    sys.stdout = sys_logger
    # Set progress text
    if progress_callback_text:
        progress_callback_text(f'Creating folder structure.')
    # Create folder structure
    create_folder_structure('./Example')
    create_folder_structure('./Example/Filled')
    create_folder_structure('./Query')
    create_folder_structure('./Query/AR1')
    create_folder_structure('./Query/AR2')
    create_folder_structure('./Log')
    create_folder_structure('./df')
    create_folder_structure('./temp')
    create_folder_structure(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}')
    logger.write(f'\nPreparing NBP_Report for:\nyear: {check_quarter()[1]},\nquarter: {check_quarter()[3]}.')
    logger.write(f'\nNBP report automation {check_quarter()[1]}')

    wb_data = create_ar2_excel()

    # Create new AR2 excel file
    sheets = wb_data[0]
    wb = wb_data[1]

    # AR2 sheet for NBP
    print(sheets[0][0])
    s1 = sheets[0][1]
    df1 = pd.DataFrame(s1.values)

    df1_insert = [
        ["D2.1", d1],
        ["D2.2", d2],
        ["D2.3", d3],
        ["D2.4", d4],
        ["D3.1", d1],
        ["D3.2", d2],
        ["D3.3", d3],
        ["D3.4", d4],

    ]

    for insert in df1_insert:
        row = df1[0][df1[0] == insert[0]].index[0]
        col = 6
        # Insert values to AR2 sheet
        s1[reference(col, 'col') + reference(row, 'row')] = insert[1]

    user = 'PAYTEL\\' + d1 + ' ' + d2
    passw = d_pass

    # Fill sheets in AR2
    prepare_data_ar2(user, passw, sheets, progress_callback, progress_callback_text)

    # Save everything to new excel file
    wb.save(wb_data[2])

    if progress_callback_text:
        progress_callback_text(f'AR2 - finished.')
    # Set progress to 50% when AR2 is completed
    if progress_callback:
        progress_callback(50)

    df_fraud_st7 = pd.read_csv(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/df_f.csv')

    # Fill sheets in AR1
    prepare_data_ar1(user, passw, df_fraud_st7, d1, d2, d3, d4, progress_callback, progress_callback_text)

    if progress_callback_text:
        progress_callback_text(f'AR1 finished.')
    # Set progress to 100% when everything is completed
    if progress_callback:
        progress_callback(100)

    # Prepare bug report tables
    # bug_report()

    # Close the log file
    log_file.close()


def start(name, surname, telephone, email, passw):
    """
    Start of main function.

    :param name: Name of person which is responsible for providing the report for NBP.
    :param surname: Surname of person which is responsible for providing the report for NBP.
    :param telephone: Telephone number of person which is responsible for providing the report for NBP.
    :param email: Email address of person which is responsible for providing the report for NBP.
    :param passw: Password of the personal account at Paytel.
    :return: 4 files. 2 xml files and 2 xlsx files.
    """
    try:
        # Read the last execution total time from "time.txt"
        last_time = 4700
        if os.path.exists("time.txt"):
            with open("time.txt", "r") as file:
                last_time = float(file.read().strip())

        d_name = name
        d_surname = surname
        d_telephone = telephone
        d_email = email
        d_pass = passw

        # Check if input is provided
        if any(input_var.strip() == '' for input_var in [d_name, d_surname, d_telephone, d_email, d_pass]):
            print('\nInput not provided. Exiting the program.')
        else:
            # Create a progress bar queue
            bar_queue = mp.Queue()

            # Start the progress bar process
            bar_process = mp.Process(target=update_bar, args=(bar_queue, last_time * 10))
            bar_process.start()

            # Run the main function in the main process
            start_automation(d_name, d_surname, d_telephone, d_email, d_pass)

            # Signal the progress bar process to finish
            bar_queue.put(None)
            bar_process.join()

    except (ValueError, TypeError, IndexError, KeyError, AttributeError, ZeroDivisionError, IOError) as e:
        # Print the error message to the console and add it to the log file
        sys.stdout.flush()  # Make sure the error message is flushed immediately
        print('\n' + f'{e}')
        raise e  # Re-raise the exception to stop the execution


if __name__ == '__main__':
    start("Krzysztof", "Kaniewski", "+48 506 297 621", "krzysztof.kaniewski@paytel.pl", 'Xl2Km0oPYahPagh8')
