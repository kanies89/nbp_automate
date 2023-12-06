import sys
import datetime
from PyQt5.QtWidgets import QApplication, QDialog, QLineEdit, QFileDialog
from PyQt5.QtCore import pyqtSignal, Qt, QEvent, QThread, pyqtSlot
from PyQt5.QtGui import QColor
from PyQt5.uic import loadUi
from main_v2 import start_automation, Logger, check_quarter
from check import AR2_TO_CHECK, AR1_TO_CHECK
import importlib
import pandas as pd
from openpyxl.utils import get_column_letter
from check_rules import check_rules_ar2, check_rules_ar1


class CheckThread(QThread):
    # Define custom signals to communicate with the main thread

    def __init__(self):
        super(CheckThread, self).__init__()

    def run(self):
        # Run the start_automation function
        print('Checking')

        # Call the start_automation function with progress_callback
        def progress_callback(step):
            self.progress = (step * 100) // self.total_steps
            self.progress_updated.emit(self.progress)

        perform_tests()

        # Emit the finished signal to indicate the completion
        self.finished.emit()

        # Enable the toolButton
        dialog.toolButton.setEnabled(True)
        dialog.radioButton_check.setEnabled(True)
        dialog.radioButton_prepare.setEnabled(True)
        dialog.pushButton_openAR1.setEnabled(True)
        dialog.pushButton_openAR2.setEnabled(True)


class AutomationThread(QThread):
    # Define custom signals to communicate with the main thread
    progress_updated = pyqtSignal(int)
    finished = pyqtSignal()
    log_updated = pyqtSignal(str)  # Custom signal to send log messages
    progress_text_updated = pyqtSignal(str)  # Custom signal for updating PasswordText label

    def __init__(self, name, surname, phone, email, password):
        super(AutomationThread, self).__init__()
        self.name = name
        self.surname = surname
        self.phone = phone
        self.email = email
        self.password = password
        self.progress = 0
        self.total_steps = 100  # You can adjust this value based on the total steps in start_automation

    def run(self):
        # Run the start_automation function
        print('starting the function')

        # Call the start_automation function with progress_callback
        def progress_callback(step):
            self.progress = (step * 100) // self.total_steps
            self.progress_updated.emit(self.progress)

        # Call the start_automation function with progress_callback_text
        def progress_callback_text(text):
            self.progress_text_updated.emit(text)  # Emit the signal to update the PasswordText label

        start_automation(self.name, self.surname, self.phone, self.email, self.password, progress_callback, progress_callback_text)

        # Emit the finished signal to indicate the completion
        self.finished.emit()

        # Enable the toolButton
        dialog.toolButton.setEnabled(True)


class MyDialog(QDialog):
    editing_finished = pyqtSignal()
    # Custom signal for progress updates
    progress_updated = pyqtSignal(int)

    @pyqtSlot(str)
    def update_progress_text(self, text):
        # Update the PasswordText label with the provided message
        self.progressLabel.setText(text)

    @pyqtSlot(str)
    def update_display(self, message):
        self.TDisplay.append(message)
        # Scroll to the last row
        scrollbar = self.TDisplay.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def __init__(self):
        super(MyDialog, self).__init__()
        # Load the UI from the XML file
        loadUi("./UI/nbp_ui.ui", self)

        self.automation_thread = None
        self.check_thread = None
        self.AR1 = None
        self.AR2 = None

        self.setup_table()  # Call the method to set up the table
        self.name = ""
        self.surname = ""
        self.enlarged = False  # track window state
        self.current_tab = 0

        # Set the fixed size of the window
        self.setFixedSize(402, 555)

        # Connect the "Apply" buttons click events to their functions
        self.BName.clicked.connect(self.on_name_apply_clicked)
        self.BSurname.clicked.connect(self.on_surname_apply_clicked)
        self.BPhone.clicked.connect(self.on_phone_apply_clicked)
        self.BEmail.clicked.connect(self.on_email_apply_clicked)
        self.Start.clicked.connect(self.on_start_clicked)

        # Connect radio buttons
        self.radioButton_check.clicked.connect(lambda button: self.radio_buttons(button='check'))
        self.radioButton_prepare.clicked.connect(lambda button: self.radio_buttons(button='prepare'))

        self.pushButton_openAR1.clicked.connect(lambda source: self.open_file_dialog(source='AR1'))
        self.pushButton_openAR2.clicked.connect(lambda source: self.open_file_dialog(source='AR2'))

        # Connect tabs
        self.tabWidget.tabBarClicked.connect(lambda index: self.toggle_window_size(index, source='tabWidget'))

        # Set the password field to display asterisks
        self.TPassword.setEchoMode(QLineEdit.Password)

        # Install the event filter for the QTextEdit widgets
        self.TName.installEventFilter(self)
        self.TSurname.installEventFilter(self)
        self.TPhone.installEventFilter(self)
        self.TEmail.installEventFilter(self)
        self.TPassword.installEventFilter(self)

        # Dictionary to map each widget with its corresponding button and the next widget in line
        self.widget_button_map = {
            self.TName: (self.BName, self.TSurname),
            self.TSurname: (self.BSurname, self.TPhone),
            self.TPhone: (self.BPhone, self.TEmail),
            self.TEmail: (self.BEmail, self.TPassword),
            self.TPassword: (self.Start, None),
        }

        # Connect the clicked signal of the toolbutton
        self.toolButton.clicked.connect(lambda: self.toggle_window_size(self.current_tab, source='toolButton'))

        # Create the logger object
        report_date = datetime.datetime.now().strftime("%Y-%m-%d")
        log_file_name = f'Log/{report_date}_LOG.txt'
        log_file = open(log_file_name, "w")
        self.logger = Logger(log_file)

        # Connect the log_updated signal from the logger to the update_display slot
        self.logger.log_updated.connect(self.update_display)

        # Assign the logger as the new sys.stdout
        sys.stdout = self.logger

    def save_logs(self):
        report_date = datetime.datetime.now().strftime("%Y-%m-%d")
        log_file_name = f'Log/{report_date}_LOG.txt'
        with open(log_file_name, 'w') as log_file:
            log_file.write(self.TDisplay.toPlainText())

    # Method to change the background color of a cell
    def change_cell_background(self, row, col, r, g, b, excel):
        if excel == 'AR2':
            item = self.tableWidget_AR2.item(row, col)
        if excel == 'AR1-ST.01':
            item = self.tableWidget_AR1_1.item(row, col)
        if excel == 'AR1-ST.02':
            item = self.tableWidget_AR1_2.item(row, col)
        if excel == 'AR1-ST.03':
            item = self.tableWidget_AR1_3.item(row, col)
        if excel == 'AR1-ST.04':
            item = self.tableWidget_AR1_4.item(row, col)
        if excel == 'AR1-ST.05':
            item = self.tableWidget_AR1_5.item(row, col)
        if excel == 'AR1-ST.06':
            item = self.tableWidget_AR1_6.item(row, col)
        if excel == 'AR1-ST.07':
            item = self.tableWidget_AR1_7.item(row, col)
        if item:
            item.setBackground(QColor(r, g, b))

    def setup_table(self):
        # Set default alignment for wrapped text in header
        for col in range(self.tableWidget_AR2.columnCount()):
            header_item = self.tableWidget_AR2.horizontalHeaderItem(col)
            if header_item:
                header_item.setTextAlignment(Qt.AlignHCenter | Qt.TextWordWrap)

        # Enable word wrapping for all cells
        for row in range(self.tableWidget_AR2.rowCount()):
            for col in range(self.tableWidget_AR2.columnCount()):
                item = self.tableWidget_AR2.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignVCenter | Qt.TextWordWrap)
                    self.tableWidget_AR2.setItem(row, col, item)
                    if col > 1:  # Adjust the column index where alignment should be set
                        item.setTextAlignment(
                            Qt.AlignVCenter | Qt.AlignHCenter)  # Align center for columns 2 and onward

        # Connect cell content change signal to row height adjustment
        self.tableWidget_AR2.cellChanged.connect(self.adjust_row_heights)

        # Set the width and height of the QtableWidget_AR2
        self.tableWidget_AR2.setFixedSize(1572, 500)  # Adjust the values as needed
        self.tableWidget_AR2.setColumnWidth(0, 80)
        self.tableWidget_AR2.setColumnWidth(1, 200)
        self.tableWidget_AR2.setColumnWidth(3, 50)
        self.tableWidget_AR2.setColumnWidth(4, 50)
        self.tableWidget_AR2.setColumnWidth(6, 95)
        self.tableWidget_AR2.setColumnWidth(10, 100)
        self.tableWidget_AR2.setColumnWidth(11, 105)
        self.tableWidget_AR2.setColumnWidth(12, 90)
        self.tableWidget_AR2.setColumnWidth(13, 50)
        self.tableWidget_AR2.setColumnWidth(14, 60)
        self.tableWidget_AR2.setColumnWidth(15, 70)
        self.tableWidget_AR2.setColumnWidth(16, 70)

        for i in range(1, 8):
            exec(f"""
# Set default alignment for wrapped text in header
for col in range(self.tableWidget_AR1_{i}.columnCount()):
    header_item = self.tableWidget_AR1_{i}.horizontalHeaderItem(col)
    if header_item:
        header_item.setTextAlignment(Qt.AlignHCenter | Qt.TextWordWrap)

# Enable word wrapping for all cells
for row in range(self.tableWidget_AR1_{i}.rowCount()):
    for col in range(self.tableWidget_AR2.columnCount()):
        item = self.tableWidget_AR1_{i}.item(row, col)
        if item:
            item.setTextAlignment(Qt.AlignVCenter | Qt.TextWordWrap)
            self.tableWidget_AR1_{i}.setItem(row, col, item)
            if col > 1:  # Adjust the column index where alignment should be set
                item.setTextAlignment(
                    Qt.AlignVCenter | Qt.AlignHCenter)  # Align center for columns 2 and onward
            
self.tableWidget_AR1_{i}.setFixedSize(1072, 500)  # Adjust the values as needed
self.tableWidget_AR1_{i}.setColumnWidth(0, 80)
self.tableWidget_AR1_{i}.setColumnWidth(1, 200)
self.tableWidget_AR1_{i}.setColumnWidth(1, 350)

# Connect cell content change signal to row height adjustment
self.tableWidget_AR1_{i}.cellChanged.connect(self.adjust_row_heights)
""")

    def adjust_row_heights(self, row, col):
        # Adjust the row height based on the contents of the specified cell
        self.tableWidget_AR2.resizeRowToContents(row)

    def toggle_window_size(self, index, source):
        if self.enlarged and source == "toolButton":
            self.setFixedSize(402, 555)  # Set your original size
            self.enlarged = False
        else:
            if index == 0:  # tab_AR2
                self.setFixedSize(1975, 555)  # Set the enlarged size
                self.current_tab = 0
                self.enlarged = True
            elif index == 1:  # tab_AR1
                self.setFixedSize(1075, 555)  # Set the enlarged size
                self.enlarged = True
                self.current_tab = 1

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress and obj in [self.TName, self.TSurname, self.TPhone, self.TEmail,
                                                       self.TPassword]:
            if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
                # Get the corresponding button and next widget
                button, next_widget = self.widget_button_map[obj]

                # Emit the editing_finished signal and click the appropriate button
                self.editing_finished.emit()
                button.click()

                # Move the focus to the next widget in line if available
                if next_widget:
                    next_widget.setFocus()

                return True

        return super().eventFilter(obj, event)

    def on_name_apply_clicked(self):
        name = self.TName.toPlainText()
        # Do something with the name input
        print("Name:", name)

        # Store the name
        self.name = name

        # Disable the current input field and enable the next one
        self.TName.setEnabled(False)
        self.BName.setEnabled(False)
        self.TSurname.setEnabled(True)
        self.BSurname.setEnabled(True)

    def on_surname_apply_clicked(self):
        surname = self.TSurname.toPlainText()
        # Do something with the surname input
        print("Surname:", surname)

        # Store the surname
        self.surname = surname

        # Disable the current input field and enable the next one
        self.TSurname.setEnabled(False)
        self.BSurname.setEnabled(False)
        self.TPhone.setEnabled(True)
        self.BPhone.setEnabled(True)

    def on_phone_apply_clicked(self):
        phone = self.TPhone.toPlainText()

        # Disable the current input field and enable the next one
        self.TPhone.setEnabled(False)
        self.BPhone.setEnabled(False)
        self.TEmail.setEnabled(True)
        self.BEmail.setEnabled(True)

    def on_email_apply_clicked(self):
        email = self.TEmail.toPlainText()

        # Disable the current input field and enable the next one
        self.TEmail.setEnabled(False)
        self.BEmail.setEnabled(False)
        self.TPassword.setEnabled(True)
        self.Start.setEnabled(True)

        # Update the PasswordText label with the provided email
        user_text = f"Provide password to your personal account for user: {self.name} {self.surname}"
        self.PasswordText.setText(user_text)

    def on_start_clicked(self):
        if self.Start.text() == 'Start':
            password = self.TPassword.text()  # Use text() method instead of toPlainText()

            self.Start.text() == 'Exit'

            # Disable the current input field and enable the "Start" button
            self.TPassword.setEnabled(False)
            self.Start.setEnabled(False)
            self.progressBar.setEnabled(True)

            # Get the input values from other fields
            name = self.TName.toPlainText()  # Use toPlainText() for QTextEdit
            surname = self.TSurname.toPlainText()  # Use toPlainText() for QTextEdit
            phone = self.TPhone.toPlainText()  # Use toPlainText() for QTextEdit
            email = self.TEmail.toPlainText()  # Use toPlainText() for QTextEdit

            # Create the AutomationThread and start it
            self.automation_thread = AutomationThread(name, surname, phone, email, password)

            # Connect the password_text_updated signal from the AutomationThread to the update_password_text slot
            self.automation_thread.progress_text_updated.connect(self.update_progress_text)
            self.automation_thread.progress_updated.connect(self.update_progress)
            self.automation_thread.finished.connect(self.on_automation_finished)

            # Connect the log_updated signal to the logger.log_updated signal
            self.automation_thread.log_updated.connect(self.logger.log_updated)

            self.pushButton_openAR1.setEnabled(False)
            self.pushButton_openAR2.setEnabled(False)
            self.radioButton_prepare.setEnabled(False)
            self.radioButton_check.setEnabled(False)

            # Start the automation thread
            self.automation_thread.start()

        elif self.Start.text() == 'Exit':
            QApplication.quit()

    def update_progress(self, progress):
        # Update the progress bar
        self.progressBar.setValue(progress)

    def append_text_to_cell(self, row, col, text_to_append, excel):
        if excel == 'AR2':
            current_item = self.tableWidget_AR2.item(row, col)
        if excel == 'AR1-ST.01':
            current_item = self.tableWidget_AR1_1.item(row, col)
        if excel == 'AR1-ST.02':
            current_item = self.tableWidget_AR1_2.item(row, col)
        if excel == 'AR1-ST.03':
            current_item = self.tableWidget_AR1_3.item(row, col)
        if excel == 'AR1-ST.04':
            current_item = self.tableWidget_AR1_4.item(row, col)
        if excel == 'AR1-ST.05':
            current_item = self.tableWidget_AR1_5.item(row, col)
        if excel == 'AR1-ST.06':
            current_item = self.tableWidget_AR1_6.item(row, col)
        if excel == 'AR1-ST.07':
            current_item = self.tableWidget_AR1_7.item(row, col)
        if current_item:
            current_text = current_item.text()
            new_text = current_text + " " + text_to_append
            current_item.setText(new_text)

    def on_automation_finished(self):

        # Update the PasswordText label with the provided email
        progress_text = f"Report finished. Checking border conditions..."
        self.progressLabel.setText(progress_text)

        self.pushButton_openAR1.setEnabled(False)
        self.pushButton_openAR2.setEnabled(False)
        self.radioButton_prepare.setEnabled(False)
        self.radioButton_check.setEnabled(False)
        # Create the AutomationThread and start it
        self.check_thread = CheckThread()
        # self.automation_thread.progress_updated.connect(self.update_progress)
        # self.automation_thread.finished.connect(self.on_automation_finished)

        # Start the check thread
        self.check_thread.start()

        # Enable the "Start" button when the automation is finished
        self.Start.setEnabled(True)
        self.Start.setText("Exit")
        # Save the logs to the log file
        self.save_logs()

    def radio_buttons(self, button):
        if button == 'prepare':
            self.radioButton_check.setChecked(False)
            self.TName.setEnabled(True)
            self.BName.setEnabled(True)
            self.pushButton_openAR1.setEnabled(False)
            self.pushButton_openAR2.setEnabled(False)
            self.AR1 = None
            self.AR2 = None
            self.toolButton.setEnabled(False)
            self.enlarged = True
            self.toggle_window_size(self.current_tab, source="toolButton")

        if button == 'check':
            self.radioButton_prepare.setChecked(False)
            self.TName.setEnabled(False)
            self.BName.setEnabled(False)
            self.pushButton_openAR1.setEnabled(True)
            self.pushButton_openAR2.setEnabled(True)
            self.enlarged = True
            self.toggle_window_size(self.current_tab, source="toolButton")

    def open_file_dialog(self, source):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*)", options=options)

        if file_name:
            if source == 'AR1':
                self.AR1 = file_name
            if source == 'AR2':
                self.AR2 = file_name

        if self.AR1 is not None and self.AR2 is not None:
            self.pushButton_openAR1.setEnabled(False)
            self.pushButton_openAR2.setEnabled(False)
            self.radioButton_prepare.setEnabled(False)
            self.radioButton_check.setEnabled(False)
            # Create the AutomationThread and start it
            self.check_thread = CheckThread()
            # self.automation_thread.progress_updated.connect(self.update_progress)
            # self.automation_thread.finished.connect(self.on_automation_finished)

            # Start the check thread
            self.check_thread.start()


def run_rule_ar1(col, dataframe, rule_number, row, sheet):
    rule_module = importlib.import_module('check')
    rule_function_name = f'rule_{rule_number}_ar1'
    rule_function = getattr(rule_module, rule_function_name, None)

    if rule_function is None or not callable(rule_function):
        raise ValueError(f"Rule {rule_function_name} not found or not callable")

    bool = True
    results = rule_function(dataframe, sheet - 1)

    for result in results:
        if not result[1]:
            dialog.change_cell_background(row, col, 255, 0, 0, f'AR1-ST.0{sheet}')
            dialog.append_text_to_cell(row, col, f'; Error in column: {result[2]}; ', f'AR1-ST.0{sheet}')
            bool = False
        if result[1] and bool:
            dialog.change_cell_background(row, col, 50, 205, 50, f'AR1-ST.0{sheet}')


def run_rule(ar, df):
    # [sheet, boolean, row, column]
    rules_2 = ["PCP_090", "PCP_091", "PCP_092", "PCP_093", "PCP_094", "PCP_096", "PCP_099", "PCP_006",
               "PCP_095", "PCP_102", "PCP_105", "PCP_007", "PCP_108", "PCP_120", "PCP_109", "PCP_121",
               "PCP_111", "PCP_123", "PCP_245_R", "DSDs_038_R", "DSDs_040_R", "PCP_110", "PCP_122"]
    rules_1 = ["RW_ST.05_01", "RW_ST.05_02"
               # , "RW_ST.05_03", "RW_ST.05_04", "RW_ST.05_05", "RW_ST.05_06",
               # "RW_ST.05_07", "RW_ST.05_08", "RW_ST.05_09", "RW_ST.05_10", "RW_ST.05_11", "RW_ST.05_12",
               # "RW_ST.05_13"
               ]

    rule: str

    if ar == 2:
        for rule in rules_2:
            results = check_rules_ar2(ar, df, rule, 6)

            for result in results:
                row = result[2]
                n = result[0]
                if result[1]:
                    dialog.change_cell_background(row, n + 3, 50, 205, 50, 'AR2')
                else:
                    dialog.change_cell_background(row, n + 3, 255, 0, 0, 'AR2')
                    dialog.append_text_to_cell(row, n + 3, f'; Error in column: {get_column_letter(result[3] + 1)}; ',
                                               'AR2')
    elif ar == 1:
        for rule in rules_1:
            results = check_rules_ar1(ar, df, rule, 13)
            if results is None:
                continue
            for result in results:
                row = result[2]
                col = 3
                sheet = result[0]

                if result[1]:
                    dialog.change_cell_background(row, col, 50, 205, 50, f'AR1-ST.0{sheet}')
                else:
                    dialog.change_cell_background(row, col, 255, 0, 0, f'AR1-ST.0{sheet}')
                    dialog.append_text_to_cell(row, col, f'; Error in column: {get_column_letter(result[3] + 1)}; ',
                                               'AR1')


def perform_tests():
    date = check_quarter()
    if dialog.AR2 is None:
        path_2 = f'EXAMPLE\\Filled\\' + f'{check_quarter()[1]}_Q{check_quarter()[3]}___BSP_AR2_v.4.01.xlsx'
    else:
        path_2 = dialog.AR2

    if dialog.AR1 is None:
        path_1 = f'EXAMPLE\\Filled\\' + f'{check_quarter()[1]}_Q{check_quarter()[3]}___BSP_AR1_ST.w.8.7.5.xlsx'
    else:
        path_1 = dialog.AR1

    # Perform all the tests
    df_nbp_2 = pd.read_excel(path_2, sheet_name=AR2_TO_CHECK, header=None, keep_default_na=False)
    df_nbp_1 = pd.read_excel(path_1, sheet_name=AR1_TO_CHECK, header=None, keep_default_na=False)
    run_rule(2, df_nbp_2)
    run_rule(1, df_nbp_1)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dialog = MyDialog()
    dialog.show()
    sys.exit(app.exec_())
