import sys
import re
import os

import datetime
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
import openpyxl

from tqdm import tqdm
import time
import multiprocessing as mp

from connect import connect
from variables import EXCEL_READ_AR2, TO_FILL, AR2_4_row_1, AR2_4_row_2, AR2_6_row_1, AR2_6_row_2, AR2_5_row_1, \
    AR2_5_row_2, \
    AR2_9_row_1, EXCEL_READ_AR1
from f_visa import f_visa_make, check_quarter, read_remote_file
from f_mastercard import f_mastercard_make
from PyQt5.QtCore import pyqtSignal, QObject

bug_table = []

TEMP = f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/'

path = 'Example\\'
df_nbp_2 = pd.read_excel(path + 'BSP_AR2_v.4.0_Q12023_20230421.xlsx', sheet_name=EXCEL_READ_AR2, header=None,
                         keep_default_na=False)
df_nbp_1 = pd.read_excel(path + 'AR1 - Q1.2023.xlsx', sheet_name=EXCEL_READ_AR1, header=None, keep_default_na=False)


class Logger(QObject):
    # Define a custom signal for log updates
    log_updated = pyqtSignal(str)

    def __init__(self, log_file):
        super(Logger, self).__init__()
        self.terminal = sys.stdout
        self.log_file = log_file

    def write(self, message):
        self.terminal.write(message)
        self.log_file.write(message)
        self.log_updated.emit(message)

    def flush(self):
        self.terminal.flush()
        self.log_file.flush()


def measure_time_with_progress(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()  # Start time
        result = func(*args, **kwargs)
        elapsed_time = time.time() - start_time  # Elapsed time
        with open("time.txt", "w") as file:
            file.write(str(elapsed_time))  # Save elapsed time to file
        print(f"\n Elapsed time: {elapsed_time:.2f}s")
        return result

    return wrapper


def progress_bar_with_elapsed_time(data):
    with tqdm(total=len(data), ncols=80, bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}') as pbar:
        for _ in data:
            time.sleep(0.1)  # Simulate some processing time
            pbar.update(1)


def to_log():
    report_date = datetime.datetime.now().strftime("%Y-%m-%d")
    file_name = f'Log/{report_date}_LOG.txt'
    with open(file_name, 'a') as log:
        log.write(f"\nReport start -AR2- --{report_date}--\n")
    return file_name


def copy_wb(from_workbook, to_workbook, dataframe, number):
    # Load the existing workbook
    wb = openpyxl.load_workbook(from_workbook)

    if number == 1:
        for sheet_name in EXCEL_READ_AR1:
            sheet = wb[sheet_name]
            for row in dataframe[sheet_name].index:
                for col in dataframe[sheet_name].columns:
                    coord = openpyxl.utils.get_column_letter(col + 1) + str(row + 1)
                    new_value = dataframe[sheet_name].iat[row, col]

                    # Handle merged cells
                    for merged_range in sheet.merged_cells.ranges:
                        if coord in merged_range:
                            # Find the first cell in the merged range
                            first_cell = merged_range.min_row, merged_range.min_col
                            first_coord = openpyxl.utils.get_column_letter(first_cell[1]) + str(first_cell[0])
                            # wb[sheet_name][first_coord].value = new_value

                            # Merge the cells
                            # sheet.merge_cells(merged_range.coord)

                            break  # Exit the loop after setting the value and merging cells
                    else:
                        # If the cell is not merged, set the value directly
                        sheet[coord].value = new_value

    elif number == 2:
        for sheet_name in EXCEL_READ_AR2:
            sheet = wb[sheet_name]
            for row in dataframe[sheet_name].index:
                for col in dataframe[sheet_name].columns:
                    coord = openpyxl.utils.get_column_letter(col + 1) + str(row + 1)
                    new_value = dataframe[sheet_name].iat[row, col]

                    # Handle merged cells
                    for merged_range in sheet.merged_cells.ranges:
                        if coord in merged_range:
                            # Find the first cell in the merged range
                            first_cell = merged_range.min_row, merged_range.min_col
                            first_coord = openpyxl.utils.get_column_letter(first_cell[1]) + str(first_cell[0])
                            # wb[sheet_name][first_coord].value = new_value

                            # Merge the cells
                            # sheet.merge_cells(merged_range.coord)

                            break  # Exit the loop after setting the value and merging cells
                    else:
                        # If the cell is not merged, set the value directly
                        sheet[coord].value = new_value

    # Save the changes
    wb.save(to_workbook)
    return wb


def load_df(length, path):
    df = []
    for n in range(length):
        df.append(pd.read_csv(f'{path}{n}.csv', keep_default_na=False))
    return df


def load_or_query(length, name, temp_table, query):
    for i in range(length):
        if os.path.exists(f'{TEMP}{name}{i}.csv'):
            if i == (length - 1) and os.path.exists(f'{TEMP}{name}{i}.csv'):
                dataframe = load_df(length, f'{TEMP}{name}')
                break
            continue
        else:
            dataframe = connect(temp_table, query)
            i = 0
            for df in dataframe:
                df.to_csv(f'{TEMP}{name}{i}.csv')
                i += 1
            break
    return dataframe


def prepare_data_ar2(user, passw, progress_callback=None):
    # 4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW
    print('4.a.R.L_PLiW2 and 4a.R.W_PLiW2 and 6.ab.LiW')

    temp_table = f"Query\\AR2\\NBP_Temp_1.sql"
    query = f"Query\\AR2\\NBP_Query_1.sql"

    dataframe_1 = load_or_query(22, '4.a.R.L_PLiW2_4a.R.W_PLiW2_6.ab.LiW__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(15)

    for n in range(len(dataframe_1) - 2):
        for i, country in enumerate(dataframe_1[n]['name']):
            if country == 'Holandia':
                country = 'Niderlandy'
            try:
                col1 = pd.Index(df_nbp_2['4a.R.L_PLiW2'].iloc[7]).get_loc(country)
                col2 = pd.Index(df_nbp_2['4a.R.W_PLiW2'].iloc[7]).get_loc(country)
                df_nbp_2['4a.R.L_PLiW2'][col1].iloc[AR2_4_row_1[n]] = dataframe_1[n]['ilosc'].iloc[i]
                df_nbp_2['4a.R.W_PLiW2'][col2].iloc[AR2_4_row_2[n]] = dataframe_1[n]['wartosc'].iloc[i]
            except KeyError:
                bug_table.append([f'BUG_4a.R.L_PLiW2', dataframe_1[n][dataframe_1[n]['name'] == country]])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_1[n][dataframe_1[n]['name'] == country]}")

    # 6.ab.LiW
    print('6.ab.LiW')

    sheet = EXCEL_READ_AR2[6]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(19)

    for j in range(2):
        df_nbp_2[sheet][33].iloc[AR2_6_row_1[j]] = dataframe_1[20 + j]['ilosc'].iloc[0]
        df_nbp_2[sheet][33].iloc[AR2_6_row_2[j]] = dataframe_1[20 + j]['wartosc'].iloc[0]

    # 5a.R.SF
    print('5a.R.SF')

    temp_table = f"Query\\AR2\\NBP_Temp_3.sql"
    query = f"Query\\AR2\\NBP_Query_3.sql"

    dataframe_3 = load_or_query(18, '5a.R.SF__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(27)

    sheet = '5a.R.SF'

    df_nbp_2[sheet][3].iloc[8] = dataframe_3[2]['wartosc'].iloc[0]
    df_nbp_2[sheet][3].iloc[9] = dataframe_3[3]['wartosc'].iloc[0]
    df_nbp_2[sheet][3].iloc[10] = dataframe_3[4]['wartosc'].iloc[0]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(29)

    # 5a.R.LF_PLiW2 and 5a.R.WF_PLiW2
    print('5a.R.LF_PLiW2 and 5a.R.WF_PLiW2')

    # Get the Visa
    data_visa = f_visa_make(user, passw)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(35)

    df_visa = pd.DataFrame(data_visa[0])
    # Get the Mastercard
    data_mastercard = f_mastercard_make()

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(39)

    df_mastercard = pd.DataFrame(data_mastercard[0])
    df_complete = [df_visa, df_mastercard]
    df_fraud = pd.concat(df_complete)
    # Use extracted ARNs for query
    with open("./Query/AR2/NBP_Temp_2.sql", 'r',
              encoding='utf-8') as sql:
        sql = sql.read()
        pattern = '--insert'
        insert_start = [match.start() for match in re.finditer(pattern, sql)]
        mastercard_insert = [insert_start[0], insert_start[0] + len(pattern)]
        visa_insert = [insert_start[1], insert_start[1] + len(pattern)]

        sql = sql[:visa_insert[0]] + data_visa[1] + sql[visa_insert[1]:]
        sql = sql[:mastercard_insert[0]] + data_mastercard[1] + sql[mastercard_insert[1]:]

    with open("./Query/AR2/NBP_Temp_2_filled.sql", 'w',
              encoding='utf-8') as sql_w:
        sql_w.write(sql)

    temp_table = f"Query\\AR2\\NBP_Temp_2_filled.sql"
    query = f"Query\\AR2\\NBP_Query_2.sql"

    dataframe_2 = load_or_query(18, '5a.R.LF_PLiW2_5a.R.WF_PLiW2__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(43)

    sheet = '5a.R.LF_PLiW2'

    j = 0
    i = 0
    for n in range(0, len(dataframe_2)):
        for country in dataframe_2[n]['code']:
            try:
                col = pd.Index(df_nbp_2[sheet].iloc[6]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', dataframe_2[dataframe_2['code'] == country]])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_2[dataframe_2['code'] == country]}")
            df_nbp_2[sheet][col].iloc[AR2_5_row_1[j]] = dataframe_2[n]['ilosc'].iloc[i]
            i += 1
        i = 0
        j += 1

    sheet = '5a.R.WF_PLiW2'

    j = 0
    i = 0
    for n in range(0, len(dataframe_2)):
        for country in dataframe_2[n]['code']:
            try:
                col = pd.Index(df_nbp_2[sheet].iloc[6]).get_loc(country)
            except KeyError:
                bug_table.append([f'BUG_{sheet}', dataframe_2[dataframe_2['code'] == country]])
                print(
                    f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_2[dataframe_2['code'] == country]}")

            df_nbp_2[sheet][col].iloc[AR2_5_row_2[j]] = dataframe_2[n]['wartosc'].iloc[i]
            i += 1
        i = 0
        j += 1

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(47)

    # Make pivot table
    df_fraud['country_aggr'] = df_fraud['country'].apply(lambda c: aggr_country(c))

    path_df = f'.\\temp\\{check_quarter()[1]}_{check_quarter()[3]}\\'
    df_fraud.to_csv(path_df + 'df_fraud.csv')

    print('\nChecking the quarter: ' + str(check_quarter()[3]))

    # 9.R.L.MCC and 9.R.W.MCC
    print('9.R.L.MCC and 9.R.W.MCC')

    temp_table = f"Query\\AR2\\NBP_Temp_4.sql"
    query = f"Query\\AR2\\NBP_Query_4.sql"

    dataframe_4 = load_or_query(4, '9.R.L.MCC_9.R.W.MCC__', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(49)

    sheet = '9.R.L.MCC'

    i = 0
    j = 0
    for n in range(0, len(dataframe_4)):
        for country in dataframe_4[n]['name']:
            # last dataframe retrieved from database is different so if n=3 then execute different algorithm
            if n < 3:
                try:
                    col = pd.Index(df_nbp_2[sheet].iloc[7]).get_loc(country)
                    df_nbp_2[sheet][col].iloc[AR2_9_row_1[n]] = dataframe_4[n]['ilosc'].iloc[i]
                except KeyError:
                    bug_table.append([f'BUG_{sheet}', dataframe_4[dataframe_4['name'] == country]])
                    print(
                        f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_4[dataframe_4['name'] == country]}")

            if n == 3:
                # Convert df_nbp_2[1] column to string
                df_nbp_2[sheet][1] = df_nbp_2[sheet][1].astype(str)
                mcc = dataframe_4[n]['mcc'].iloc[i]
                try:
                    col = pd.Index(df_nbp_2[sheet].iloc[7]).get_loc(country)
                    ind = df_nbp_2[sheet][df_nbp_2[sheet][1] == mcc].index[0]
                    df_nbp_2[sheet].iat[ind, col] = dataframe_4[n]['ilosc'].iloc[i]
                except KeyError:
                    bug_table.append([f'BUG_{sheet}', dataframe_4[dataframe_4['name'] == country]])
                    print(
                        f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_4[dataframe_4['name'] == country]}")

            i += 1
        i = 0
        j += 1

    sheet = '9.R.W.MCC'

    i = 0
    j = 0
    for n in range(0, len(dataframe_4)):
        for country in dataframe_4[n]['name']:
            # last dataframe retrieved from database is different so if n=3 then execute different algorithm
            if n < 3:
                try:
                    col = pd.Index(df_nbp_2[sheet].iloc[7]).get_loc(country)
                    df_nbp_2[sheet][col].iloc[AR2_9_row_1[n]] = dataframe_4[n]['wartosc_transakcji'].iloc[i]
                except KeyError:
                    bug_table.append([f'BUG_{sheet}', dataframe_4[dataframe_4['name'] == country]])
                    print(
                        f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_4[dataframe_4['name'] == country]}")

            if n == 3:
                # Convert df_nbp_2[1] column to string
                df_nbp_2[sheet][1] = df_nbp_2[sheet][1].astype(str)
                mcc = dataframe_4[n]['mcc'].iloc[i]
                try:
                    col = pd.Index(df_nbp_2[sheet].iloc[7]).get_loc(country)
                    ind = df_nbp_2[sheet][df_nbp_2[sheet][1] == mcc].index[0]
                    df_nbp_2[sheet].iat[ind, col] = dataframe_4[n]['wartosc_transakcji'].iloc[i]
                except KeyError:
                    bug_table.append([f'BUG_{sheet}', dataframe_4[dataframe_4['name'] == country]])
                    print(
                        f"!!: Value was not added to the report (there is no such a country code in excel) - {dataframe_4[dataframe_4['name'] == country]}")

            i += 1
        i = 0
        j += 1

    return df_fraud


def aggr_country(c):
    if c == 'PL':
        return 'PL'
    else:
        return 'NPL'


def prepare_data_ar1(user, passw, df_f, name, surname, phone, email, progress_callback=None):
    # ST.01
    print('ST.01')

    temp_table = f"Query\\AR1\\NBP_Temp_1.sql"
    query = f"Query\\AR1\\NBP_Query_1.sql"

    dataframe_1 = load_or_query(2, 'ST.01.', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(55)

    # Get data from "Tvid_nev_lost.xlsx'
    path = f'//prdfil/Business/DPiUS/Zespol Przetwarzania/Raporty kwartalne/{check_quarter()[1]}Q{check_quarter()[3]}/Tvid_nev_lost.xlsx'
    read_remote_file(path, user, passw)
    dataframe_0 = pd.read_excel(path, header=3)

    print('\nData z pliku Tvid_nev_lost.xlsx: ', dataframe_0['tr_date'][0])

    # Data to be filled
    to_change_values = [
        dataframe_1[1]['MID all cashback'][0],
        dataframe_1[1]['SID all cashback'][0],
        dataframe_1[1]['TVID all cashback'][0],
        dataframe_1[0]['ilosc_softPOS'][0],
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_sid'][0]  # sid all

    ]
    to_change_rows = [8, 13, 18, 21, 15, 16, 17, 19, 6, 7, 11, 12]
    to_change_column = 5

    # Changes in dataframe from spreadsheet - df_nbp_1
    for v in range(len(to_change_values)):
        df_nbp_1['ST.01'].iat[to_change_rows[v], to_change_column] = to_change_values[v]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(57)

    # ST.03
    print('ST.03')

    to_change_values = [
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_mid'][0],  # mid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_sid'][0],  # sid all
        dataframe_0['active_tvid'][0],  # tvid all
        dataframe_0['active_tvid'][0]  # tvid all

    ]
    to_change_rows = [7, 8, 12, 13, 16, 17]
    to_change_column = 5

    # Just using values filled to sheet ST.01
    for v in range(len(to_change_values)):
        df_nbp_1['ST.03'].iat[to_change_rows[v], to_change_column] = to_change_values[v]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(64)

    # ST.05
    print('ST.05')

    temp_table = f"Query\\AR1\\NBP_Temp_2.sql"
    query = f"Query\\AR1\\NBP_Query_2.sql"

    dataframe_2 = load_or_query(6, 'ST.05.', temp_table, query)

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(68)

    # Data to be filled
    def prepare_values_data(d, c):
        if d == 1:
            ilosc = 'Liczba transakcji CashBack'
            wartosc = 'Wartość wypłat Cash Back'
        elif d == 2:
            ilosc = 'Liczba transakcji BLIK'
            wartosc = 'Kwota transakcji BLIK'
        else:
            ilosc = 'ilosc'
            wartosc = 'wartosc'

        if d == 2:
            to_change_values = [
                dataframe_2[d][ilosc][0],
                0,
                dataframe_2[d][wartosc][0],
                0,
                dataframe_2[d][ilosc][0],
                0,
                dataframe_2[d][wartosc][0],
                0
            ]
        else:
            to_change_values = [
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    ilosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'POLSKA') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum(),
                dataframe_2[d][(dataframe_2[d]['KRAJE'] == 'INNE KRAJE') & (dataframe_2[d]['category'] == c)][
                    wartosc].sum()
            ]
        return to_change_values

    to_change_columns = [9, 10, 11, 12, 13, 14, 15, 16]
    to_change_rows_1 = [13, 11, 12, 27, 21]
    to_change_rows_2 = [17, 15, 16, 29, 21]
    df = [1, 3, 4, 5, 2]

    # Changes in dataframe from spreadsheet - df_nbp_1

    for row in range(len(df)):
        category = 'Individual'
        to_change_values = prepare_values_data(df[row], category)
        for v in range(len(to_change_values)):
            df_nbp_1['ST.05'].iat[to_change_rows_1[row], to_change_columns[v]] = to_change_values[v]

    for row in range(len(df)):
        category = 'BUSINESS'
        to_change_values = prepare_values_data(df[row], category)
        for v in range(len(to_change_values)):
            df_nbp_1['ST.05'].iat[to_change_rows_2[row], to_change_columns[v]] = to_change_values[v]

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(75)

    # ST.06
    print('ST.06')

    temp_table = f"Query\\AR1\\NBP_Temp_3.sql"
    query = f"Query\\AR1\\NBP_Query_3.sql"

    dataframe_3 = load_or_query(2, 'ST.06.', temp_table, query)
    print('\nDate: ' + str(dataframe_3[0]))

    dataframe_3 = dataframe_3[1]

    column_amount = [3, 4, 5]
    column_value = [6, 7, 8]

    content_amount = ['ilosc_transakcji', 'ilosc_internet', 'ilosc_transakcji_CashBack']
    content_value = ['wartosc_transakcji', 'wartosc_internet', 'wartosc_wyplat_CashBack']

    geo6 = pd.read_excel("C:\\Users\\Krzysztof kaniewski\\PycharmProjects\\pythonProject\\Example\\NBP_GEO6.xlsx",
                         header=3)

    # devices that accept payment cards / Internet / cash back
    for k, country in enumerate(dataframe_3['CountryCode']):
        if country == 'uwaga - coś nowego':
            print(
                f"!!: In the report fraud transactions were not added (probably due to NULL value) - {dataframe_3[dataframe_3['CountryCode'] == 'uwaga - coś nowego']}")
            bug_table.append([f'ST.06', dataframe_3[dataframe_3['CountryCode'] == 'uwaga - coś nowego']])
        elif country in df_nbp_1['ST.06'][0][10:39].values:
            row = pd.Index(df_nbp_1['ST.06'][0]).get_loc(country)
            for i in range(len(column_amount)):
                df_nbp_1['ST.06'].iat[row, column_amount[i]] = dataframe_3[dataframe_3['CountryCode'] == country][
                    content_amount[i]].values[0]
                df_nbp_1['ST.06'].iat[row, column_value[i]] = dataframe_3[dataframe_3['CountryCode'] == country][
                    content_value[i]].values[0]
        else:
            try:
                if country == 'NA':
                    c_name = 'Namibia'
                    new_row_data = ['NA', c_name, c_name] + [np.nan] * (len(column_amount) + len(column_value))
                    for i in range(len(column_amount)):
                        new_row_data[column_amount[i]] = dataframe_3[dataframe_3['name_PL'] == 'Namibia'][content_amount[i]].values[0]
                        new_row_data[column_value[i]] = dataframe_3[dataframe_3['name_PL'] == 'Namibia'][content_value[i]].values[0]
                else:
                    country_name = geo6[geo6['Code'] == country]['Nazwa kraju'].values[0]
                    country_name_english = geo6[geo6['Code'] == country]['Name'].values[0]
                    new_row_data = [country, country_name, country_name_english] + [np.nan] * (len(column_amount) + len(column_value))
                    for i in range(len(column_amount)):
                        new_row_data[column_amount[i]] = dataframe_3[dataframe_3['CountryCode'] == country][content_amount[i]].values[0]
                        new_row_data[column_value[i]] = dataframe_3[dataframe_3['CountryCode'] == country][content_value[i]].values[0]
                if country != 'GB':
                    # Concatenate new row data to DataFrame
                    new_row_data_t = pd.DataFrame(new_row_data).T
                    new_row_data_t.reset_index(drop=True, inplace=True)
                    df_nbp_1['ST.06'] = pd.concat([df_nbp_1['ST.06'], new_row_data_t], axis=0)
                else:
                    new_row_data = pd.DataFrame([new_row_data])
                    # Concatenate the original DataFrame with the new row DataFrame
                    df_nbp_1['ST.06'] = pd.concat([df_nbp_1['ST.06'].iloc[:40], new_row_data, df_nbp_1['ST.06'].iloc[40:]], ignore_index=True)
                df_nbp_1['ST.06'].reset_index(drop=True, inplace=True)
            except KeyError:
                print(
                    f"!!: In the report fraud transactions were not added (there is no such a country code in excel) - {dataframe_3[dataframe_3['CountryCode'] == country]}")
                bug_table.append([f'ST.06', dataframe_3[dataframe_3['CountryCode'] == country]])

    for i in range(len(column_amount)):
        # Convert the numeric string values to actual numeric types for sum calculation
        df_nbp_1['ST.06'][column_amount[i]][10:] = pd.to_numeric(df_nbp_1['ST.06'][column_amount[i]][10:])
        df_nbp_1['ST.06'][column_value[i]][10:] = pd.to_numeric(df_nbp_1['ST.06'][column_value[i]][10:])

        # Assign the sum values to row 39
        df_nbp_1['ST.06'][column_amount[i]].loc[39] = df_nbp_1['ST.06'][column_amount[i]][40:].sum()
        df_nbp_1['ST.06'][column_value[i]].loc[39] = df_nbp_1['ST.06'][column_value[i]][40:].sum()

    # Calculate and update progress to 30% when appropriate
    if progress_callback:
        progress_callback(86)

    # ST.02

    # Right now we do not fill this sheet.

    # ST.07
    print('ST.07')

    temp_table = f"Query\\AR1\\NBP_Temp_4.sql"
    query = f"Query\\AR1\\NBP_Query_4.sql"

    dataframe_4 = load_or_query(4, 'ST.07.', temp_table, query)
    # Calculate and update progress to 30% when appropriate

    if progress_callback:
        progress_callback(98)

    dataframe_4[1]['kwota'] = dataframe_4[1]['kwota'].astype('float')
    df_nbp_1['ST.07'].iat[14, 4] = dataframe_4[1][dataframe_4[1]['kraj'] == 'PL']['ilosc'].iloc[0]
    df_nbp_1['ST.07'].iat[14, 5] = dataframe_4[1][dataframe_4[1]['kraj'] == 'other']['ilosc'].iloc[0]
    df_nbp_1['ST.07'].iat[14, 6] = dataframe_4[1][dataframe_4[1]['kraj'] == 'PL']['kwota'].iloc[0]
    df_nbp_1['ST.07'].iat[14, 7] = dataframe_4[1][dataframe_4[1]['kraj'] == 'other']['kwota'].iloc[0]
    df_nbp_1['ST.07'].iat[14, 8] = dataframe_4[3]['kwota'].iloc[0]
    df_nbp_1['ST.07'].iat[14, 9] = dataframe_4[2]['kwota'].iloc[0]

    dff = pd.pivot_table(index='pos_entry_mode', columns='country_aggr',
                         data=df_f[df_f['quarter'] == check_quarter()[3]],
                         aggfunc={'tr_amout': 'sum', 'country_aggr': 'count'}, fill_value=0)

    df_nbp_1['ST.07'].iat[11, 4] = dff['country_aggr'].sum()['PL']
    df_nbp_1['ST.07'].iat[11, 5] = dff['country_aggr'].sum()['NPL']
    df_nbp_1['ST.07'].iat[11, 6] = dff['tr_amout'].sum()['PL']
    df_nbp_1['ST.07'].iat[11, 7] = dff['tr_amout'].sum()['NPL']
    df_nbp_1['ST.07'].iat[11, 8] = 0
    df_nbp_1['ST.07'].iat[11, 9] = 0

    df_nbp_1['ST.07'].iat[12, 4] = dff['country_aggr'].loc['CTLS']['PL']
    df_nbp_1['ST.07'].iat[12, 5] = dff['country_aggr'].loc['CTLS']['NPL']
    df_nbp_1['ST.07'].iat[12, 6] = dff['tr_amout'].loc['CTLS']['PL']
    df_nbp_1['ST.07'].iat[12, 7] = dff['tr_amout'].loc['CTLS']['NPL']
    df_nbp_1['ST.07'].iat[12, 8] = 0
    df_nbp_1['ST.07'].iat[12, 9] = 0

    for n in range(4, 10):
        df_nbp_1['ST.07'].iat[10, n] = float(df_nbp_1['ST.07'][n].iloc[11:15].sum()) - float(
            df_nbp_1['ST.07'][n].iloc[12])

    for n in range(4, 10):
        df_nbp_1['ST.07'].iat[9, n] = df_nbp_1['ST.07'].iat[10, n]


    author_data = [
        name,
        surname,
        phone,
        email
    ]

    for x in range(2, 9):
        i = 0
        for y in range(8, 12):
            df_nbp_1['p-dane'].iat[y, x] = author_data[i]
            i += 1
    print('END')


def bug_report():
    wb_sheet_names = []
    if not bug_table:
        print('Full success - no bugs!')
    else:
        wb_bug = pd.ExcelWriter(f'{TEMP}bug_table.xlsx', engine='openpyxl')

        for i, bug in enumerate(bug_table):
            if bug[0] + '_0' not in wb_sheet_names:
                sheet_name = bug[0] + '_0'
                wb_sheet_names.append(sheet_name)
                bug[1].to_excel(wb_bug, sheet_name=sheet_name, index=False)
                print(f'Added bug table to sheet: {sheet_name}')
            else:
                no = []
                for sheet_name in wb_sheet_names:
                    if bug[0] + '_' in sheet_name:
                        try:
                            no.append(int(sheet_name[len(bug[0] + '_'):]))
                        except IndexError:
                            print('Error in writing the bug table.')

                max_no = max(no) + 1
                sheet_name = bug[0] + '_' + str(max_no)
                wb_sheet_names.append(sheet_name)
                bug[1].to_excel(wb_bug, sheet_name=sheet_name, index=False)
                print(f'Added bug table to sheet: {sheet_name}')

        # Add all the bug tables to a single sheet
        combined_sheet_name = 'Combined Sheet'
        all_bugs = pd.DataFrame(columns=['Sheet Name', 'Data', 'Sum'])

        for sheet_name in wb_sheet_names:
            data = bug_table[wb_sheet_names.index(sheet_name)][1]
            if 'wartosc_transakcji' in data.columns and 'wartosc_internet' in data.columns and 'wartosc_wyplat_CashBack' in data.columns:
                data_sum = data['wartosc_transakcji'] + data['wartosc_internet'] + data['wartosc_wyplat_CashBack']
                data_sum = data_sum.item() if len(data_sum) == 1 else None  # Extract the float value from Series
            else:
                data_sum = None

            row_data = {'Sheet Name': sheet_name, 'Data': data, 'Sum': data_sum}
            all_bugs = pd.concat([all_bugs, pd.DataFrame([row_data])], ignore_index=True)

        # Reorder the sheets
        writer = pd.ExcelWriter(f'{TEMP}bug_table.xlsx', engine='openpyxl')
        writer.book = wb_bug.book
        all_bugs['Sum'] = all_bugs['Sum'].astype(float)  # Convert the 'Sum' column to float
        all_bugs.to_excel(writer, sheet_name=combined_sheet_name, index=False)
        writer.save()

        print(f'Added all bug tables to the combined sheet: {combined_sheet_name}')

        wb_bug.close()
        print('Saved bug tables to bug_table.xlsx')


@measure_time_with_progress
def start_automation(d1, d2, d3, d4, d_pass, progress_callback=None):
    # Open the log file in append mode
    log_file = open(to_log(), "a")

    # Create the logger object
    logger = Logger(log_file)

    # Create an instance of the Logger class
    sys_logger = Logger(sys.stdout)

    # Assign the logger as the new sys.stdout
    sys.stdout = sys_logger

    # Create folder structure
    create_folder_structure('./Example')
    create_folder_structure('./Example/Filled')
    create_folder_structure('./Query')
    create_folder_structure('./Query/AR1')
    create_folder_structure('./Query/AR2')
    create_folder_structure('./Log')
    create_folder_structure('./df')
    create_folder_structure('./temp')
    create_folder_structure(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}')
    logger.write(f'\nPreparing NBP_Report for:\nyear: {check_quarter()[1]},\nquarter: {check_quarter()[3]}.')
    logger.write(f'\nNBP report automation {check_quarter()[1]}')

    # AR2 sheet for NBP
    # Fill the first sheet with "Author of the report" info.
    # input the personal data
    d_21 = d1
    d_22 = d2
    d_23 = d3
    d_24 = d4
    d_31 = d_21
    d_32 = d_22
    d_33 = d_23
    d_34 = d_24
    user = 'PAYTEL\\' + d_21 + ' ' + d_22
    passw = d_pass

    input_data = [
        d_21, d_22, d_23, d_24, d_31, d_32, d_33, d_34
    ]

    # RETURN ROW WITH "D2.1" IN COLUMN 0 - COLUMN 5 to be edited
    i = 0
    for inp in input_data:
        df_nbp_2[EXCEL_READ_AR2[0]].loc[df_nbp_2[EXCEL_READ_AR2[0]][0] == TO_FILL[i], 5] = inp
        i += 1

    # Fill sheets in AR2
    df_fraud_st7 = prepare_data_ar2(user, passw, progress_callback)
    df_fraud_st7.to_csv(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/df_f.csv')

    # Save everything to new excel file
    from_wb = path + 'BSP_AR2_v.4.0_Q12023_20230421.xlsx'
    to_wb = path + f'Filled\\' + f'BSP_AR2_v.4.0_Q{check_quarter()[3]}{datetime.date.today().strftime("%Y")}_{datetime.date.today().strftime("%Y%m%d")}.xlsx'

    wb = copy_wb(from_wb, to_wb, df_nbp_2, 2)

    # Save the updated workbook
    wb.save(to_wb)

    # Set progress to 50% when AR2 is completed
    if progress_callback:
        progress_callback(50)

    df_fraud_st7 = pd.read_csv(f'./temp/{check_quarter()[1]}_{check_quarter()[3]}/df_f.csv')
    # Fill sheets in AR1
    prepare_data_ar1(user, passw, df_fraud_st7, d_21, d_22, d_23, d_24, progress_callback)

    # Save everything to new excel file
    from_wb = path + 'AR1 - Q1.2023.xlsx'
    to_wb = path + f'Filled\\' + f'AR1 - Q{check_quarter()[3]}.{datetime.date.today().strftime("%Y")}.xlsx'

    wb = copy_wb(from_wb, to_wb, df_nbp_1, 1)

    # Save the updated workbook
    wb.save(to_wb)

    # Set progress to 100% when everything is completed
    if progress_callback:
        progress_callback(100)

    # Prepare bug report tables
    bug_report()

    # Close the log file
    log_file.close()


class TqdmExtraFormat(tqdm):
    """Provides a `minutes per iteration` format parameter"""

    @property
    def format_dict(self):
        d = super(TqdmExtraFormat, self).format_dict
        rate_min = '{:.2f}'.format(1 / d["rate"] / 60) if d["rate"] else '?'
        d.update(rate_min=(rate_min + ' min/' + d['unit']))
        return d


def create_folder_structure(address):
    # Check if the folder structure exists
    if not os.path.exists(address):
        # Create the folder structure
        os.makedirs(address)
        print("Folder structure created successfully.")
    else:
        print("Folder structure already exists.")


def update_bar(queue, total):
    pbar = tqdm(total=round(total, 2))

    while True:
        if pbar.n >= pbar.total:
            break
        time.sleep(0.1)  # Simulate some processing time
        pbar.update(1)
        elapsed_time = (time.time() - pbar.start_t)
        pbar.set_postfix(elapsed=f"{elapsed_time:.2f}s",
                         remaining=f"{(pbar.total - pbar.n) * (elapsed_time / pbar.n):.2f}s / {((pbar.total - pbar.n) * (elapsed_time / pbar.n)) / 60:.2f}min")

    pbar.close()


def start(name, surname, telephone, email, passw):
    try:
        # Read the last execution total time from "time.txt"
        last_time = 4700
        if os.path.exists("time.txt"):
            with open("time.txt", "r") as file:
                last_time = float(file.read().strip())

        d_name = name
        d_surname = surname
        d_telephone = telephone
        d_email = email
        d_pass = passw

        # Check if input is provided
        if any(input_var.strip() == '' for input_var in [d_name, d_surname, d_telephone, d_email, d_pass]):
            print('\nInput not provided. Exiting the program.')
        else:
            # Create a progress bar queue
            bar_queue = mp.Queue()

            # Start the progress bar process
            bar_process = mp.Process(target=update_bar, args=(bar_queue, last_time * 10))
            bar_process.start()

            # Run the main function in the main process
            start_automation(d_name, d_surname, d_telephone, d_email, d_pass)

            # Signal the progress bar process to finish
            bar_queue.put(None)
            bar_process.join()

    except (ValueError, TypeError, IndexError, KeyError, AttributeError, ZeroDivisionError, IOError) as e:
        # Print the error message to the console and add it to the log file
        sys.stdout.flush()  # Make sure the error message is flushed immediately
        print('\n' + f'{e}')
        raise e  # Re-raise the exception to stop the execution


if __name__ == '__main__':
    start()